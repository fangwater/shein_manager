from __future__ import annotations

import base64
import csv
import hashlib
import hmac
import io
import json
import os
import re
import secrets
import subprocess
import sys
import threading
import time
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from html import escape
from pathlib import Path
from typing import Any

import pandas as pd
import psycopg
from psycopg.rows import dict_row

from fastapi import Body, Cookie, FastAPI, File, Form, HTTPException, Query, UploadFile

from .config import load_settings
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response

BASE_DIR = Path(__file__).resolve().parents[2]
TEMPLATE_DIR = Path(__file__).resolve().parent / "web_templates"
ITEMS_PATH = BASE_DIR / "exports" / "shein_order_items_profit.parquet"
COOKIE_NAME = "shein_pnl_session"
LEGACY_COOKIE_NAME = "shein_pnl_token"
COOKIE_SECURE = os.getenv("SHEIN_WEB_COOKIE_SECURE", "true").strip().lower() not in {"0", "false", "no"}
SESSION_TTL_SECONDS = 7 * 24 * 3600
SESSION_SECRET_PATH = Path(os.getenv("SHEIN_WEB_SESSION_SECRET_FILE", BASE_DIR / ".web_session_secret"))
ALL_PERMISSIONS = "*"
VIEW_PROFIT = "view_profit"
VIEW_RETURNS = "view_returns"
VIEW_LOGISTICS = "view_logistics"
VIEW_SHIPPING_FEE = "view_shipping_fee"
ACCESS_SKU_MAPPINGS = "access_sku_mappings"
VIEW_WAREHOUSE_RELATIONS = "view_warehouse_relations"
VIEW_WAREHOUSE_COST = "view_warehouse_cost"
ACCESS_INVENTORY = "access_inventory"
ACCESS_COST_TEMPLATES = "access_cost_templates"
ROLE_ADMIN = "admin"
ROLE_TEST = "test"
ROLE_OPERATIONS = "operations"
ROLE_ORDER_FOLLOW_UP = "order_follow_up"
ROLE_PERMISSIONS: dict[str, frozenset[str]] = {
    ROLE_ADMIN: frozenset({ALL_PERMISSIONS}),
    ROLE_TEST: frozenset({
        VIEW_RETURNS,
        VIEW_LOGISTICS,
        VIEW_SHIPPING_FEE,
        ACCESS_SKU_MAPPINGS,
        VIEW_WAREHOUSE_RELATIONS,
        VIEW_WAREHOUSE_COST,
        ACCESS_INVENTORY,
        ACCESS_COST_TEMPLATES,
    }),
    ROLE_OPERATIONS: frozenset({
        VIEW_RETURNS,
        VIEW_LOGISTICS,
        VIEW_SHIPPING_FEE,
        ACCESS_SKU_MAPPINGS,
        VIEW_WAREHOUSE_RELATIONS,
    }),
    ROLE_ORDER_FOLLOW_UP: frozenset({
        VIEW_RETURNS,
        ACCESS_SKU_MAPPINGS,
        VIEW_WAREHOUSE_RELATIONS,
        VIEW_WAREHOUSE_COST,
        ACCESS_INVENTORY,
        ACCESS_COST_TEMPLATES,
    }),
}
ROLE_HOME_PATHS = {
    ROLE_ADMIN: "/",
    ROLE_TEST: "/sku-mappings",
    ROLE_OPERATIONS: "/logistics",
    ROLE_ORDER_FOLLOW_UP: "/sku-mappings",
}
APP_TITLE = "Panda SHEIN PNL"
ACTUAL_PNL_MODE = "actual"
ESTIMATED_PNL_MODE = "estimated"
NAV_PERMISSION_HREFS = {
    VIEW_PROFIT: ("./",),
    VIEW_LOGISTICS: ("logistics",),
    VIEW_SHIPPING_FEE: ("shipping-fee",),
    VIEW_RETURNS: ("returns",),
    ACCESS_SKU_MAPPINGS: ("sku-mappings",),
    VIEW_WAREHOUSE_RELATIONS: ("warehouse-relations",),
    ACCESS_INVENTORY: ("inventory",),
    ACCESS_COST_TEMPLATES: ("inventory-templates",),
}
ESTIMATE_METHOD_ZERO = "zero"
ESTIMATE_METHOD_SKU_AVG = "sku_avg"
RETURN_PROFIT_ZEROED_POLICY = "return_profit_zeroed"
REVENUE_EXCLUDED_POLICY = "revenue_excluded_status"
PENDING_PICKUP_WITHOUT_FULFILLMENT_POLICY = "pending_pickup_without_fulfillment_fee"
PENDING_PICKUP_ESTIMATED_POLICY = "pending_pickup_estimated"
PENDING_PICKUP_ESTIMATE_MISSING_POLICY = "pending_pickup_estimate_missing"
AFTER_SALES_POLICY = "after_sales_cost"
REVENUE_EXCLUDED_POLICIES = {REVENUE_EXCLUDED_POLICY, PENDING_PICKUP_WITHOUT_FULFILLMENT_POLICY, PENDING_PICKUP_ESTIMATE_MISSING_POLICY}
SALES_EXCLUDED_POLICIES = {RETURN_PROFIT_ZEROED_POLICY, *REVENUE_EXCLUDED_POLICIES}
LOGISTICS_EXCLUDED_STATUS_CODES = {"1", "2", "7"}
LOGISTICS_EXCLUDED_STATUS_NORMALIZED = {"pending", "pending_shipment", "pending_pickup"}
LOGISTICS_EXCLUDED_STATUS_PATTERN = r"待揽收|待发货|待处理|未确认|未发货"
SHIPPING_FEE_DIMENSIONS = [
    {"value": "all", "label": "整体"},
    {"value": "warehouse_sku", "label": "仓库 SKU"},
    {"value": "shein_sku", "label": "SHEIN SKU"},
    {"value": "skc", "label": "SKC"},
    {"value": "sku_attr", "label": "商品规格"},
    {"value": "country", "label": "国家"},
    {"value": "province", "label": "省州"},
]
SHIPPING_FEE_DEFAULT_DIMENSION = "warehouse_sku"
SHIPPING_FEE_DEFAULT_ROLLING_DAYS = 7
SHIPPING_FEE_DEFAULT_MIN_PERIODS = SHIPPING_FEE_DEFAULT_ROLLING_DAYS
SHIPPING_FEE_DEFAULT_SHIFT_DAYS = 7
SHIPPING_FEE_DEFAULT_TOP_N = 12

app = FastAPI(title=APP_TITLE)
REFRESH_EXPORT_LOCK = threading.Lock()
SYNC_LATEST_ORDERS_LOCK = threading.Lock()
SYNC_LATEST_RETURNS_LOCK = threading.Lock()
SESSION_SECRET_LOCK = threading.Lock()
_SESSION_SECRET: bytes | None = None


@dataclass(frozen=True)
class Account:
    username: str
    password: str
    role: str

    @property
    def permissions(self) -> frozenset[str]:
        return ROLE_PERMISSIONS[self.role]

    @property
    def home_path(self) -> str:
        return ROLE_HOME_PATHS[self.role]

    def can(self, permission: str) -> bool:
        return ALL_PERMISSIONS in self.permissions or permission in self.permissions


def load_accounts() -> dict[str, Account]:
    return {
        "pyy": Account(
            username="pyy",
            password=os.getenv("SHEIN_WEB_PYY_PASSWORD", "12345"),
            role=ROLE_ADMIN,
        ),
        "temu-test": Account(
            username="temu-test",
            password=os.getenv("SHEIN_WEB_TEMU_TEST_PASSWORD", "temu-test"),
            role=ROLE_TEST,
        ),
        "operations": Account(
            username="operations",
            password=os.getenv("SHEIN_WEB_OPERATIONS_PASSWORD", "operations"),
            role=ROLE_OPERATIONS,
        ),
        "order-follow-up": Account(
            username="order-follow-up",
            password=os.getenv("SHEIN_WEB_ORDER_FOLLOW_UP_PASSWORD", "order-follow-up"),
            role=ROLE_ORDER_FOLLOW_UP,
        ),
    }


def session_secret() -> bytes:
    global _SESSION_SECRET
    if _SESSION_SECRET is not None:
        return _SESSION_SECRET
    with SESSION_SECRET_LOCK:
        if _SESSION_SECRET is not None:
            return _SESSION_SECRET
        configured = os.getenv("SHEIN_WEB_SESSION_SECRET", "").strip()
        if configured:
            _SESSION_SECRET = configured.encode("utf-8")
            return _SESSION_SECRET
        try:
            value = SESSION_SECRET_PATH.read_text(encoding="utf-8").strip()
        except FileNotFoundError:
            value = secrets.token_urlsafe(48)
            SESSION_SECRET_PATH.parent.mkdir(parents=True, exist_ok=True)
            try:
                fd = os.open(SESSION_SECRET_PATH, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600)
            except FileExistsError:
                value = SESSION_SECRET_PATH.read_text(encoding="utf-8").strip()
            else:
                with os.fdopen(fd, "w", encoding="utf-8") as handle:
                    handle.write(value)
        if not value:
            raise RuntimeError("SHEIN Web session secret is empty")
        _SESSION_SECRET = value.encode("utf-8")
        return _SESSION_SECRET


def encode_session(account: Account) -> str:
    payload = json.dumps(
        {"username": account.username, "expires": int(time.time()) + SESSION_TTL_SECONDS},
        separators=(",", ":"),
        sort_keys=True,
    ).encode("utf-8")
    encoded = base64.urlsafe_b64encode(payload).rstrip(b"=")
    signature = hmac.new(session_secret(), encoded, hashlib.sha256).hexdigest().encode("ascii")
    return (encoded + b"." + signature).decode("ascii")


def decode_session(value: str | None) -> Account | None:
    if not value:
        return None
    try:
        encoded, supplied_signature = value.encode("ascii").rsplit(b".", 1)
        expected_signature = hmac.new(session_secret(), encoded, hashlib.sha256).hexdigest().encode("ascii")
        if not hmac.compare_digest(supplied_signature, expected_signature):
            return None
        padding = b"=" * (-len(encoded) % 4)
        payload = json.loads(base64.urlsafe_b64decode(encoded + padding))
        if int(payload.get("expires", 0)) < int(time.time()):
            return None
        return load_accounts().get(str(payload.get("username", "")))
    except (ValueError, TypeError, json.JSONDecodeError, UnicodeError):
        return None


def read_template(name: str) -> str:
    return (TEMPLATE_DIR / name).read_text(encoding="utf-8")


def render_template(name: str, account: Account) -> str:
    html = read_template(name)
    for permission, hrefs in NAV_PERMISSION_HREFS.items():
        if not account.can(permission):
            href_pattern = "|".join(re.escape(href) for href in hrefs)
            html = re.sub(rf'<a\b[^>]*href="(?:{href_pattern})"[^>]*>.*?</a>', "", html)
    if name == "warehouse_relations.html" and not account.can(VIEW_WAREHOUSE_COST):
        html = html.replace("<div>成本</div><div>${esc(r.costText||'-')}</div>", "")
        html = re.sub(
            r"\$\{detailList\('成本',r\.costList,.*?(?=\$\{detailList\('库存')",
            "",
            html,
        )
    return (
        html.replace("{{username}}", escape(account.username))
        .replace("{{role}}", escape(account.role))
    )


def login_html(message: str = "") -> str:
    html = read_template("login.html")
    return html.replace("{{message}}", escape(message)).replace("{{title}}", APP_TITLE)


def page_response(template_name: str, cookie_value: str | None, *, permission: str | None = None) -> Response:
    account = decode_session(cookie_value)
    if account is None:
        return HTMLResponse(login_html())
    if permission and not account.can(permission):
        raise HTTPException(status_code=403, detail="permission denied")
    return HTMLResponse(render_template(template_name, account))


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/", response_class=HTMLResponse)
def index(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> HTMLResponse:
    account = decode_session(shein_pnl_token)
    if account is None:
        return HTMLResponse(login_html())
    if not account.can(VIEW_PROFIT):
        return RedirectResponse(account.home_path, status_code=303)
    return HTMLResponse(render_template("dashboard.html", account))


@app.get("/returns", response_class=HTMLResponse)
def returns_page(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> HTMLResponse:
    return page_response("returns.html", shein_pnl_token, permission=VIEW_RETURNS)


@app.get("/sku-mappings", response_class=HTMLResponse)
def sku_mappings_page(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> HTMLResponse:
    return page_response("sku_mappings.html", shein_pnl_token, permission=ACCESS_SKU_MAPPINGS)




@app.get("/warehouse-relations", response_class=HTMLResponse)
def warehouse_relations_page(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> HTMLResponse:
    return page_response("warehouse_relations.html", shein_pnl_token, permission=VIEW_WAREHOUSE_RELATIONS)


@app.get("/inventory", response_class=HTMLResponse)
def inventory_page(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> HTMLResponse:
    return page_response("inventory.html", shein_pnl_token, permission=ACCESS_INVENTORY)


@app.get("/inventory-templates", response_class=HTMLResponse)
def inventory_templates_page(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> HTMLResponse:
    return page_response("inventory_templates.html", shein_pnl_token, permission=ACCESS_COST_TEMPLATES)


@app.get("/logistics", response_class=HTMLResponse)
def logistics_page(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> HTMLResponse:
    return page_response("logistics.html", shein_pnl_token, permission=VIEW_LOGISTICS)


@app.get("/shipping-fee", response_class=HTMLResponse)
def shipping_fee_page(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> HTMLResponse:
    return page_response("shipping_fee.html", shein_pnl_token, permission=VIEW_SHIPPING_FEE)


@app.post("/login")
def login(username: str = Form(...), password: str = Form(...)) -> Response:
    account = load_accounts().get(username.strip())
    if account is None or not secrets.compare_digest(password, account.password):
        return HTMLResponse(login_html("用户名或密码不正确"), status_code=401)
    response = RedirectResponse(account.home_path, status_code=303)
    response.set_cookie(
        COOKIE_NAME,
        encode_session(account),
        httponly=True,
        secure=COOKIE_SECURE,
        samesite="lax",
        max_age=SESSION_TTL_SECONDS,
        path="/",
    )
    response.delete_cookie(LEGACY_COOKIE_NAME, path="/")
    return response


@app.get("/logout")
def logout() -> Response:
    response = RedirectResponse("/", status_code=303)
    response.delete_cookie(COOKIE_NAME, path="/")
    response.delete_cookie(LEGACY_COOKIE_NAME, path="/")
    return response


def require_auth(token: str | None, cookie_token: str | None, *, permission: str | None = None) -> Account:
    account = decode_session(cookie_token)
    if account is None:
        raise HTTPException(status_code=401, detail="authentication required")
    if permission and not account.can(permission):
        raise HTTPException(status_code=403, detail="permission denied")
    return account


def parse_shein_datetime(value: Any) -> pd.Series | pd.Timestamp:
    if isinstance(value, pd.Series):
        text = value.astype("string").str.strip()
        text = text.str.replace(r"([+-]\d{2}:?\d{2}|Z)$", "", regex=True)
        return pd.to_datetime(text, errors="coerce", format="mixed")
    if value in (None, ""):
        return pd.NaT
    text = str(value).strip()
    text = re.sub(r"([+-]\d{2}:?\d{2}|Z)$", "", text)
    return pd.to_datetime(text, errors="coerce")


def load_items() -> pd.DataFrame:
    if not ITEMS_PATH.exists():
        raise HTTPException(status_code=500, detail=f"missing export: {ITEMS_PATH}")
    df = pd.read_parquet(ITEMS_PATH).copy()
    df["order_created_dt"] = parse_shein_datetime(df["order_created_at"])
    df["sku_label"] = df["sku_attr_us"].fillna("").astype(str).replace({"": "未匹配"})
    for col, fallback in {"pnl_product_cost_usd": "product_cost_rule_usd", "pnl_packaging_fee_usd": "packaging_fee_rule_usd", "pnl_internal_cost_usd": "internal_cost_usd"}.items():
        if col not in df and fallback in df:
            df[col] = df[fallback]
    if "after_sales_cost_usd" not in df:
        df["after_sales_cost_usd"] = 0.0
    if "revenue_excluded_status" not in df:
        df["revenue_excluded_status"] = False
    if "pending_pickup_without_fulfillment_fee" not in df:
        df["pending_pickup_without_fulfillment_fee"] = False
    if "after_sales_status" not in df:
        df["after_sales_status"] = False
    if "pnl_policy" not in df:
        df["pnl_policy"] = "standard"
    for original_col, active_col in {
        "original_base_revenue_allocated_usd": "base_revenue_allocated_usd",
        "original_shipping_fee_allocated_usd": "shipping_fee_allocated_usd",
        "original_performance_service_charge_allocated_usd": "performance_service_charge_allocated_usd",
        "original_gross_revenue_allocated_usd": "gross_revenue_allocated_usd",
        "original_profit_usd": "profit_usd",
    }.items():
        if original_col not in df and active_col in df:
            df[original_col] = df[active_col]
    return df


def parse_hour(value: str | None, default: pd.Timestamp) -> pd.Timestamp:
    if not value:
        return default
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        raise HTTPException(status_code=400, detail=f"invalid datetime: {value}")
    return parsed.floor("h")


def filter_items(df: pd.DataFrame, *, skus: list[str], start: str | None, end: str | None) -> pd.DataFrame:
    min_dt = df["order_created_dt"].min().floor("h") if not df.empty else pd.Timestamp.now().floor("h")
    max_dt = df["order_created_dt"].max().ceil("h") if not df.empty else pd.Timestamp.now().ceil("h")
    start_dt = parse_hour(start, min_dt)
    end_dt = parse_hour(end, max_dt)
    if end_dt < start_dt:
        raise HTTPException(status_code=400, detail="end must be >= start")
    mask = df["order_created_dt"].between(start_dt, end_dt + pd.Timedelta(hours=1) - pd.Timedelta(microseconds=1))
    if skus:
        filter_col = "warehouse_sku_key" if "warehouse_sku_key" in df.columns else "sku_label"
        mask &= df[filter_col].isin(skus)
    return df.loc[mask].copy()


def load_pnl_relation_maps() -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    settings = load_settings()
    if not settings.database_url:
        return {}, {}
    try:
        with psycopg.connect(settings.database_url, row_factory=dict_row) as conn:
            catalog = build_product_sku_catalog(conn, settings.shop_key)
            mapping_rows = conn.execute(
                """
                SELECT shein_sku, warehouse_sku, sku_group, warehouse_qty, updated_at
                FROM shein_sku_mappings
                WHERE shop_key = %s AND enabled = true
                ORDER BY updated_at DESC, id DESC
                """,
                (settings.shop_key,),
            ).fetchall()
    except Exception:
        return {}, {}
    mappings: dict[str, dict[str, Any]] = {}
    for row in mapping_rows:
        shein_sku = clean_text(row.get("shein_sku"))
        if not shein_sku or shein_sku in mappings:
            continue
        mappings[shein_sku] = {
            "warehouse_sku": clean_text(row.get("warehouse_sku")),
            "sku_group": clean_text(row.get("sku_group")),
            "warehouse_qty": number(row.get("warehouse_qty")),
            "updated_at": row.get("updated_at"),
        }
    return mappings, catalog


def enrich_pnl_items_with_relations(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    sku_source = df["sku_code"] if "sku_code" in df.columns else pd.Series("", index=df.index)
    df["shein_sku_code"] = sku_source.fillna("").astype(str).str.strip()
    mappings, catalog = load_pnl_relation_maps()

    warehouse_skus: list[str] = []
    warehouse_labels: list[str] = []
    warehouse_keys: list[str] = []
    mapping_statuses: list[str] = []
    skc_names: list[str] = []
    skc_keys: list[str] = []
    skc_labels: list[str] = []
    spu_names: list[str] = []
    product_titles: list[str] = []
    product_images: list[str] = []
    supplier_skus: list[str] = []
    supplier_codes: list[str] = []
    category_ids: list[str] = []
    skc_shelf_statuses: list[Any] = []
    sku_cost_usd_values: list[float | None] = []
    sku_base_price_usd_values: list[float | None] = []
    sku_special_price_usd_values: list[float | None] = []

    for sku_code, sku_label in zip(df["shein_sku_code"], df["sku_label"]):
        sku_code = clean_text(sku_code)
        fallback_sku = sku_code or clean_text(sku_label) or "unknown"
        mapping = mappings.get(sku_code, {}) if sku_code else {}
        product = catalog.get(sku_code, {}) if sku_code else {}
        warehouse_sku = clean_text(mapping.get("warehouse_sku"))
        is_mapped = bool(warehouse_sku)
        warehouse_skus.append(warehouse_sku)
        warehouse_labels.append(warehouse_sku if is_mapped else f"未映射 · {fallback_sku}")
        warehouse_keys.append(warehouse_sku if is_mapped else f"__missing__:{fallback_sku}")
        mapping_statuses.append("mapped" if is_mapped else "missing")

        skc_name = clean_text(product.get("skcName"))
        spu_name = clean_text(product.get("spuName"))
        skc_names.append(skc_name)
        skc_keys.append(skc_name if skc_name else f"__missing_skc__:{fallback_sku}")
        skc_labels.append(skc_name if skc_name else f"缺商品资料 · {fallback_sku}")
        spu_names.append(spu_name)
        product_titles.append(clean_text(product.get("title")))
        product_images.append(clean_text(product.get("imageUrl")))
        supplier_skus.append(clean_text(product.get("supplierSku")))
        supplier_codes.append(clean_text(product.get("supplierCode")))
        category_ids.append(clean_text(product.get("categoryId")))
        skc_shelf_statuses.append(product.get("skcShelfStatus"))
        sku_cost_usd_values.append(product_cost_usd(product.get("costList") or []))
        sku_base_price_usd_values.append(product_price_usd(product.get("priceList") or [], "basePrice"))
        sku_special_price_usd_values.append(product_price_usd(product.get("priceList") or [], "specialPrice"))

    df["warehouse_sku"] = warehouse_skus
    df["warehouse_sku_label"] = warehouse_labels
    df["warehouse_sku_key"] = warehouse_keys
    df["warehouse_mapping_status"] = mapping_statuses
    df["warehouse_mapping_missing"] = df["warehouse_mapping_status"].ne("mapped")
    df["skc_name"] = skc_names
    df["skc_key"] = skc_keys
    df["skc_label"] = skc_labels
    df["spu_name"] = spu_names
    df["product_title"] = product_titles
    df["product_image_url"] = product_images
    df["supplier_sku"] = supplier_skus
    df["supplier_code"] = supplier_codes
    df["category_id"] = category_ids
    df["skc_shelf_status"] = skc_shelf_statuses
    df["sku_cost_usd"] = sku_cost_usd_values
    df["sku_base_price_usd"] = sku_base_price_usd_values
    df["sku_special_price_usd"] = sku_special_price_usd_values
    return df


def product_cost_usd(items: Any) -> float | None:
    for item in json_list(items):
        if isinstance(item, dict) and clean_text(item.get("currency")).upper() == "USD":
            value = number(item.get("cost"))
            return value if value > 0 else None
    return None


def product_price_usd(items: Any, field: str) -> float | None:
    rows = [item for item in json_list(items) if isinstance(item, dict)]
    preferred = [
        item for item in rows
        if clean_text(item.get("currency")).upper() == "USD"
        and clean_text(item.get("site")).lower() == "shein-us"
    ]
    fallback = [item for item in rows if clean_text(item.get("currency")).upper() == "USD"]
    for item in [*preferred, *fallback]:
        value = number(item.get(field))
        if value > 0:
            return value
    return None


def optional_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        if pd.isna(value):
            return None
        return round(float(value), 2)
    except (TypeError, ValueError):
        return None


def compact_series(values: pd.Series, *, limit: int = 4, max_chars: int = 180) -> str:
    return compact_unique([clean_text(value) for value in values.tolist()], limit=limit, max_chars=max_chars)


def first_clean_series(values: pd.Series) -> str:
    for value in values.tolist():
        text = clean_text(value)
        if text:
            return text
    return ""


def nunique_nonempty(values: pd.Series) -> int:
    return len({clean_text(value) for value in values.tolist() if clean_text(value)})


def sku_detail_rows(group: pd.DataFrame) -> list[dict[str, str]]:
    details: dict[tuple[str, str], dict[str, str]] = {}
    for row in group[["shein_sku_code", "skc_name"]].itertuples(index=False, name=None):
        shein_sku = clean_text(row[0])
        if not shein_sku:
            continue
        skc = clean_text(row[1])
        details[(shein_sku, skc)] = {"sheinSku": shein_sku, "skc": skc}
    return sorted(details.values(), key=lambda item: (item["skc"], item["sheinSku"]))


def skc_warehouse_rows(group: pd.DataFrame) -> list[dict[str, Any]]:
    rows: dict[str, dict[str, Any]] = {}
    child_seen: dict[str, set[str]] = {}
    columns = [
        "warehouse_sku_key",
        "warehouse_sku_label",
        "product_image_url",
        "shein_sku_code",
        "sku_cost_usd",
        "sku_base_price_usd",
        "sku_special_price_usd",
    ]
    for key, label, image_url, shein_sku, cost_usd, base_price, special_price in group[columns].itertuples(index=False, name=None):
        warehouse_key = clean_text(key)
        warehouse_label = clean_text(label) or warehouse_key
        if not warehouse_label:
            continue
        row_key = warehouse_key or warehouse_label
        current = rows.setdefault(row_key, {"warehouseSku": warehouse_label, "imageUrl": "", "sheinSkus": []})
        child_seen.setdefault(row_key, set())
        if not current.get("imageUrl"):
            current["imageUrl"] = clean_text(image_url)
        sku_code = clean_text(shein_sku)
        if sku_code and sku_code not in child_seen[row_key]:
            child_seen[row_key].add(sku_code)
            current["sheinSkus"].append({
                "sheinSku": sku_code,
                "costUsd": optional_float(cost_usd),
                "basePriceUsd": optional_float(base_price),
                "specialPriceUsd": optional_float(special_price),
            })
    for item in rows.values():
        item["sheinSkus"] = sorted(item["sheinSkus"], key=lambda child: child.get("sheinSku") or "")
    return sorted(rows.values(), key=lambda item: item["warehouseSku"])


def missing_mapping_summary(df: pd.DataFrame) -> dict[str, int]:
    if df.empty or "warehouse_mapping_missing" not in df.columns:
        return {"missingMappingLines": 0, "missingMappingSkus": 0}
    missing = df.loc[df["warehouse_mapping_missing"]]
    return {
        "missingMappingLines": int(len(missing)),
        "missingMappingSkus": int(missing["shein_sku_code"].nunique()) if "shein_sku_code" in missing else 0,
    }


def logistics_fee_series(df: pd.DataFrame) -> pd.Series:
    if "original_performance_service_charge_allocated_usd" in df:
        source = df["original_performance_service_charge_allocated_usd"]
    elif "performance_service_charge_allocated_usd" in df:
        source = df["performance_service_charge_allocated_usd"]
    else:
        source = pd.Series(0.0, index=df.index)
    return pd.to_numeric(source, errors="coerce").fillna(0.0)


def logistics_revenue_series(df: pd.DataFrame) -> pd.Series:
    if "original_gross_revenue_allocated_usd" in df:
        source = df["original_gross_revenue_allocated_usd"]
    elif "gross_revenue_allocated_usd" in df:
        source = df["gross_revenue_allocated_usd"]
    else:
        source = pd.Series(0.0, index=df.index)
    return pd.to_numeric(source, errors="coerce").fillna(0.0)


def logistics_sku_key(code: Any, label: Any) -> str:
    code_text = str(code or "").strip()
    label_text = str(label or "").strip() or "未匹配"
    return f"{code_text}||{label_text}"


def prepare_logistics_items(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if df.empty:
        return df
    code_source = df["sku_code"] if "sku_code" in df else pd.Series("", index=df.index)
    df["logistics_sku_code"] = code_source.fillna("").astype(str).str.strip()
    df["logistics_sku_label"] = df["sku_label"].fillna("").astype(str).replace({"": "未匹配"})
    df["logistics_sku_key"] = [
        logistics_sku_key(code, label)
        for code, label in zip(df["logistics_sku_code"], df["logistics_sku_label"])
    ]
    status_code = df.get("order_status", pd.Series("", index=df.index)).fillna("").astype(str).str.strip()
    status_label = df.get("order_status_label", pd.Series("", index=df.index)).fillna("").astype(str).str.strip()
    status_normalized = df.get("order_status_normalized", pd.Series("", index=df.index)).fillna("").astype(str).str.strip()
    excluded = (
        status_code.isin(LOGISTICS_EXCLUDED_STATUS_CODES)
        | status_normalized.isin(LOGISTICS_EXCLUDED_STATUS_NORMALIZED)
        | status_label.str.contains(LOGISTICS_EXCLUDED_STATUS_PATTERN, regex=True, na=False)
    )
    return_detected = df.get("return_detected", pd.Series(False, index=df.index)).fillna(False).astype(bool)
    after_sales_status = df.get("after_sales_status", pd.Series(False, index=df.index)).fillna(False).astype(bool)
    policy = df.get("pnl_policy", pd.Series("", index=df.index)).fillna("").astype(str)
    df["logistics_fulfillment_fee_usd"] = logistics_fee_series(df)
    df["logistics_gross_revenue_usd"] = logistics_revenue_series(df)
    df["logistics_weight"] = pd.to_numeric(df.get("goods_weight", pd.Series(0.0, index=df.index)), errors="coerce").fillna(0.0)
    df["logistics_effective_line"] = ~excluded
    df["logistics_excluded_line"] = excluded
    df["logistics_after_sales_line"] = return_detected | after_sales_status | policy.eq(RETURN_PROFIT_ZEROED_POLICY)
    df["logistics_zero_fee_line"] = df["logistics_fulfillment_fee_usd"].le(0)
    return df


def filter_logistics_items(df: pd.DataFrame, *, sku_keys: list[str], start: str | None, end: str | None) -> pd.DataFrame:
    min_dt = df["order_created_dt"].min().floor("h") if not df.empty else pd.Timestamp.now().floor("h")
    max_dt = df["order_created_dt"].max().ceil("h") if not df.empty else pd.Timestamp.now().ceil("h")
    start_dt = parse_hour(start, min_dt)
    end_dt = parse_hour(end, max_dt)
    if end_dt < start_dt:
        raise HTTPException(status_code=400, detail="end must be >= start")
    mask = df["order_created_dt"].between(start_dt, end_dt + pd.Timedelta(hours=1) - pd.Timedelta(microseconds=1))
    if sku_keys:
        mask &= df["logistics_sku_key"].isin(sku_keys)
    return df.loc[mask].copy()


def logistics_fee_buckets(fees: pd.Series) -> list[dict[str, Any]]:
    clean = pd.to_numeric(fees, errors="coerce").fillna(0.0)
    if clean.empty:
        return []
    buckets: list[dict[str, Any]] = []
    zero_count = int(clean.le(0).sum())
    if zero_count:
        buckets.append({"label": "$0", "lines": zero_count})
    edges = [0, 1, 2, 3, 4, 5, 7.5, 10, 15, 20]
    for low, high in zip(edges[:-1], edges[1:]):
        count = int(clean.gt(low).where(clean.le(high), False).sum())
        if count:
            buckets.append({"label": f"${low:g}-${high:g}", "lines": count})
    high_count = int(clean.gt(edges[-1]).sum())
    if high_count:
        buckets.append({"label": f">${edges[-1]:g}", "lines": high_count})
    return buckets


def normalize_pnl_mode(value: str | None) -> str:
    value = (value or ACTUAL_PNL_MODE).strip().lower()
    if value in {"estimate", "estimated", "shadow"}:
        return ESTIMATED_PNL_MODE
    return ACTUAL_PNL_MODE


def normalize_estimate_method(value: str | None) -> str:
    value = (value or ESTIMATE_METHOD_SKU_AVG).strip().lower()
    if value in {"0", "zero", "none"}:
        return ESTIMATE_METHOD_ZERO
    return ESTIMATE_METHOD_SKU_AVG


def build_sku_fulfillment_fee_averages(df: pd.DataFrame) -> dict[str, float]:
    if df.empty or "performance_service_charge_allocated_usd" not in df or "warehouse_sku" not in df:
        return {}
    fee = pd.to_numeric(df["performance_service_charge_allocated_usd"], errors="coerce").fillna(0.0)
    source = df.loc[fee > 0, ["warehouse_sku", "performance_service_charge_allocated_usd"]].copy()
    source["warehouse_sku"] = source["warehouse_sku"].map(clean_text)
    source = source.loc[source["warehouse_sku"].ne("")]
    if source.empty:
        return {}

    grouped = source.groupby("warehouse_sku", dropna=False)["performance_service_charge_allocated_usd"].mean()
    return {clean_text(key): float(value or 0.0) for key, value in grouped.items() if clean_text(key)}


def sku_fulfillment_fee_average_series(
    df: pd.DataFrame,
    warehouse_averages: dict[str, float],
) -> pd.Series:
    if df.empty or "warehouse_sku" not in df:
        return pd.Series(0.0, index=df.index, dtype="float64")
    values = [float(warehouse_averages.get(clean_text(value), 0.0)) for value in df["warehouse_sku"]]
    return pd.Series(values, index=df.index, dtype="float64")


def estimate_pending_pickup_fulfillment_fee(df: pd.DataFrame, *, method: str) -> pd.Series:
    method = normalize_estimate_method(method)
    if method == ESTIMATE_METHOD_ZERO:
        return pd.Series(0.0, index=df.index, dtype="float64")
    return pd.to_numeric(df.get("sku_avg_fulfillment_fee_usd", pd.Series(0.0, index=df.index)), errors="coerce").fillna(0.0)


def apply_pnl_mode(
    df: pd.DataFrame,
    *,
    mode: str,
    estimate_method: str = ESTIMATE_METHOD_SKU_AVG,
    sku_fee_averages: dict[str, float] | None = None,
) -> pd.DataFrame:
    mode = normalize_pnl_mode(mode)
    estimate_method = normalize_estimate_method(estimate_method)
    df = df.copy()
    warehouse_averages = sku_fee_averages or build_sku_fulfillment_fee_averages(df)
    df["sku_avg_fulfillment_fee_usd"] = sku_fulfillment_fee_average_series(df, warehouse_averages) if not df.empty else pd.Series(dtype="float64")
    df["pnl_mode"] = mode
    df["estimate_method"] = estimate_method
    df["is_estimated_pnl"] = False
    df["pnl_estimate_missing"] = False
    df["estimated_fulfillment_fee_allocated_usd"] = 0.0
    if df.empty or mode != ESTIMATED_PNL_MODE:
        return df

    policy = df.get("pnl_policy", pd.Series("standard", index=df.index)).fillna("standard")
    mask = policy.eq(PENDING_PICKUP_WITHOUT_FULFILLMENT_POLICY)
    if not bool(mask.any()):
        return df

    if estimate_method == ESTIMATE_METHOD_SKU_AVG:
        has_warehouse_average = df["warehouse_sku"].map(lambda value: clean_text(value) in warehouse_averages)
        estimate_missing_mask = mask & ~has_warehouse_average
        estimable_mask = mask & has_warehouse_average
    else:
        estimate_missing_mask = pd.Series(False, index=df.index)
        estimable_mask = mask

    if bool(estimate_missing_mask.any()):
        df.loc[estimate_missing_mask, "pnl_estimate_missing"] = True
        df.loc[estimate_missing_mask, "pnl_policy"] = PENDING_PICKUP_ESTIMATE_MISSING_POLICY

    if not bool(estimable_mask.any()):
        return df

    estimated_fee = estimate_pending_pickup_fulfillment_fee(df.loc[estimable_mask], method=estimate_method)
    df.loc[estimable_mask, "is_estimated_pnl"] = True
    df.loc[estimable_mask, "estimated_fulfillment_fee_allocated_usd"] = estimated_fee
    df.loc[estimable_mask, "base_revenue_allocated_usd"] = df.loc[estimable_mask, "original_base_revenue_allocated_usd"]
    df.loc[estimable_mask, "shipping_fee_allocated_usd"] = df.loc[estimable_mask, "original_shipping_fee_allocated_usd"]
    df.loc[estimable_mask, "performance_service_charge_allocated_usd"] = estimated_fee
    df.loc[estimable_mask, "gross_revenue_allocated_usd"] = df.loc[estimable_mask, "original_gross_revenue_allocated_usd"]
    df.loc[estimable_mask, "pnl_product_cost_usd"] = df.loc[estimable_mask, "product_cost_rule_usd"]
    df.loc[estimable_mask, "pnl_packaging_fee_usd"] = df.loc[estimable_mask, "packaging_fee_rule_usd"]
    df.loc[estimable_mask, "pnl_internal_cost_usd"] = df.loc[estimable_mask, "pnl_product_cost_usd"] + df.loc[estimable_mask, "pnl_packaging_fee_usd"]
    df.loc[estimable_mask, "after_sales_cost_usd"] = 0.0
    df.loc[estimable_mask, "profit_usd"] = (
        df.loc[estimable_mask, "gross_revenue_allocated_usd"]
        - df.loc[estimable_mask, "performance_service_charge_allocated_usd"]
        - df.loc[estimable_mask, "pnl_product_cost_usd"]
        - df.loc[estimable_mask, "pnl_packaging_fee_usd"]
    )
    df.loc[estimable_mask, "profit_margin"] = df.loc[estimable_mask, "profit_usd"] / df.loc[estimable_mask, "gross_revenue_allocated_usd"].replace({0: pd.NA})
    df.loc[estimable_mask, "pnl_policy"] = PENDING_PICKUP_ESTIMATED_POLICY
    return df


RETURN_TIME_FIELDS = {
    "request": ("request_return_dt", "退货申请时间"),
    "updated": ("last_update_dt", "更新时间"),
    "completed": ("completed_dt", "完成时间"),
}
RETURN_PAGE_LIMIT = 1000


def database_url() -> str:
    settings = load_settings()
    if not settings.database_url:
        raise HTTPException(status_code=500, detail="DATABASE_URL is required")
    return settings.database_url


def parse_datetime(value: str | None, default: pd.Timestamp) -> pd.Timestamp:
    if not value:
        return default
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        raise HTTPException(status_code=400, detail=f"invalid datetime: {value}")
    return parsed


def parse_json_list(value: Any) -> list[Any]:
    if isinstance(value, list):
        return value
    if value in (None, ""):
        return []
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return []
        return parsed if isinstance(parsed, list) else []
    return []


def number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def clean_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value)


def clean_number(value: Any) -> float:
    return round(number(value), 2)


def compact_unique(values: list[str], *, limit: int = 4, max_chars: int = 160) -> str:
    unique: list[str] = []
    for value in values:
        value = str(value or "").strip()
        if value and value not in unique:
            unique.append(value)
    text = ", ".join(unique[:limit])
    if len(unique) > limit:
        text += f" 等{len(unique)}项"
    return text[:max_chars]


def preferred_reason(reasons: Any) -> str:
    reason_list = parse_json_list(reasons)
    for language in ("CN", "ZH", "EN", "US"):
        for item in reason_list:
            if isinstance(item, dict) and item.get("language") == language and item.get("reason"):
                return str(item["reason"])
    for item in reason_list:
        if isinstance(item, dict) and item.get("reason"):
            return str(item["reason"])
    return ""


def summarize_return_goods(value: Any) -> dict[str, Any]:
    goods = [item for item in parse_json_list(value) if isinstance(item, dict)]
    skus: list[str] = []
    titles: list[str] = []
    reasons: list[str] = []
    seller_amount = 0.0
    performance_price = 0.0
    return_expense = 0.0
    estimate_income = 0.0
    currencies: list[str] = []
    for item in goods:
        skus.append(str(item.get("skuSn") or item.get("sku") or item.get("goodsSn") or ""))
        titles.append(str(item.get("goodsTitle") or ""))
        reasons.append(preferred_reason(item.get("returnReasonList")))
        seller_amount += number(item.get("sellerCurrencyPrice"))
        performance_price += number(item.get("performancePrice"))
        return_expense += number(item.get("returnExpense"))
        estimate_income += number(item.get("estimateIncomeMoney"))
        currencies.append(str(item.get("currency") or item.get("saleCurrency") or ""))
    return {
        "goodsLines": len(goods),
        "skus": compact_unique(skus, limit=5, max_chars=180),
        "goodsTitles": compact_unique(titles, limit=2, max_chars=180),
        "reasons": compact_unique(reasons, limit=3, max_chars=180),
        "currency": compact_unique(currencies, limit=2, max_chars=20),
        "sellerAmount": round(seller_amount, 2),
        "goodsPerformancePrice": round(performance_price, 2),
        "returnExpense": round(return_expense, 2),
        "estimateIncome": round(estimate_income, 2),
    }


def load_return_orders() -> pd.DataFrame:
    columns = [
        "return_no", "order_no", "return_status", "return_status_label", "no_return_goods_sign",
        "return_order_tag_code", "site", "platform_express_no", "member_express_no",
        "express_company_name", "performance_cost", "invoice_status", "request_return_time",
        "allocate_time", "last_update_time", "seller_signed_time", "cancel_time", "completed_time",
        "check_status", "stock_mode", "receive_type", "refund_order_nos", "return_goods_info_list",
        "order_status", "order_status_label", "order_type_label", "order_created_at",
    ]
    sql = """
        SELECT
            r.return_no, r.order_no, r.return_status, r.return_status_label,
            r.no_return_goods_sign, r.return_order_tag_code, r.site,
            r.platform_express_no, r.member_express_no, r.express_company_name,
            r.performance_cost, r.invoice_status, r.request_return_time, r.allocate_time,
            r.last_update_time, r.seller_signed_time, r.cancel_time, r.completed_time,
            r.check_status, r.stock_mode, r.receive_type, r.refund_order_nos,
            r.return_goods_info_list,
            o.order_status, o.order_status_label, o.order_type_label, o.order_created_at
        FROM shein_order_returns r
        LEFT JOIN shein_orders o ON o.shop_key = r.shop_key AND o.order_no = r.order_no
        ORDER BY COALESCE(NULLIF(r.request_return_time, ''), NULLIF(r.last_update_time, ''), NULLIF(r.allocate_time, '')) DESC NULLS LAST,
                 r.return_no DESC
    """
    with psycopg.connect(database_url(), row_factory=dict_row) as conn:
        rows = conn.execute(sql).fetchall()
    df = pd.DataFrame(rows, columns=columns)
    if df.empty:
        return df
    for col in ("request_return_time", "last_update_time", "completed_time", "allocate_time", "order_created_at"):
        df[col.replace("_time", "_dt") if col.endswith("_time") else f"{col}_dt"] = parse_shein_datetime(df[col])
    df["return_filter_dt"] = df["request_return_dt"].fillna(df["last_update_dt"]).fillna(df["allocate_dt"])
    df["performance_cost"] = pd.to_numeric(df["performance_cost"], errors="coerce").fillna(0.0)
    return df


def filter_return_orders(
    df: pd.DataFrame,
    *,
    start: str | None,
    end: str | None,
    statuses: list[str],
    time_field: str,
) -> tuple[pd.DataFrame, str]:
    if df.empty:
        return df, RETURN_TIME_FIELDS.get(time_field, RETURN_TIME_FIELDS["request"])[1]
    if time_field not in RETURN_TIME_FIELDS:
        raise HTTPException(status_code=400, detail=f"invalid timeField: {time_field}")
    time_col, time_label = RETURN_TIME_FIELDS[time_field]
    series = df[time_col].fillna(df["return_filter_dt"])
    min_dt = series.min()
    max_dt = series.max()
    if pd.isna(min_dt):
        min_dt = pd.Timestamp.now().floor("d")
    if pd.isna(max_dt):
        max_dt = pd.Timestamp.now().ceil("d")
    start_dt = parse_datetime(start, min_dt)
    end_dt = parse_datetime(end, max_dt)
    if end_dt < start_dt:
        raise HTTPException(status_code=400, detail="end must be >= start")
    mask = series.between(start_dt, end_dt + pd.Timedelta(minutes=1) - pd.Timedelta(microseconds=1))
    if statuses:
        mask &= df["return_status"].fillna("").astype(str).isin(statuses)
    return df.loc[mask].copy(), time_label


def order_pnl_lookup() -> dict[str, dict[str, float]]:
    if not ITEMS_PATH.exists():
        return {}
    try:
        df = pd.read_parquet(ITEMS_PATH)
    except Exception:
        return {}
    required = {"order_no", "internal_cost_usd", "profit_usd"}
    if not required.issubset(df.columns):
        return {}
    aggregations: dict[str, tuple[str, str]] = {
        "pnlInternalCost": ("internal_cost_usd", "sum"),
        "pnlProfit": ("profit_usd", "sum"),
    }
    if "original_gross_revenue_allocated_usd" in df.columns:
        aggregations["originalGrossRevenue"] = ("original_gross_revenue_allocated_usd", "sum")
    if "original_profit_usd" in df.columns:
        aggregations["originalProfit"] = ("original_profit_usd", "sum")
    grouped = df.groupby("order_no", dropna=False).agg(**aggregations).reset_index()
    return {str(row["order_no"]): {key: round(float(row.get(key) or 0), 2) for key in aggregations} for row in grouped.to_dict(orient="records")}


def return_time_bounds(df: pd.DataFrame) -> tuple[str, str]:
    if df.empty:
        now = pd.Timestamp.now().floor("min")
        return now.strftime("%Y-%m-%dT%H:%M"), now.strftime("%Y-%m-%dT%H:%M")
    series = df["return_filter_dt"]
    min_dt = series.min()
    max_dt = series.max()
    if pd.isna(min_dt) or pd.isna(max_dt):
        now = pd.Timestamp.now().floor("min")
        return now.strftime("%Y-%m-%dT%H:%M"), now.strftime("%Y-%m-%dT%H:%M")
    return min_dt.floor("min").strftime("%Y-%m-%dT%H:%M"), max_dt.ceil("min").strftime("%Y-%m-%dT%H:%M")


def summarize(df: pd.DataFrame, *, mode: str = ACTUAL_PNL_MODE, estimate_method: str = ESTIMATE_METHOD_SKU_AVG) -> dict[str, Any]:
    estimate_method = normalize_estimate_method(estimate_method)
    if df.empty:
        return {"pnlMode": mode, "estimateMethod": estimate_method, "orders": 0, "salesOrders": 0, "returnOrders": 0, "revenueExcludedOrders": 0, "estimatedOrders": 0, "estimatedLines": 0, "estimateMissingOrders": 0, "estimateMissingLines": 0, "afterSalesOrders": 0, "lines": 0, "baseRevenue": 0, "shippingRevenue": 0, "grossRevenue": 0, "performanceFee": 0, "productCost": 0, "packagingFee": 0, "afterSalesCost": 0, "profit": 0, "margin": None}
    gross = float(df["gross_revenue_allocated_usd"].sum())
    profit = float(df["profit_usd"].sum())
    policy = df.get("pnl_policy", pd.Series("standard", index=df.index)).fillna("standard")
    estimated_mask = df.get("is_estimated_pnl", pd.Series(False, index=df.index)).fillna(False).astype(bool)
    return_mask = policy.eq(RETURN_PROFIT_ZEROED_POLICY)
    revenue_excluded_mask = policy.isin(REVENUE_EXCLUDED_POLICIES)
    after_sales_mask = policy.eq(AFTER_SALES_POLICY)
    sales_mask = ~policy.isin(SALES_EXCLUDED_POLICIES)
    sales_df = df.loc[sales_mask]
    sales_gross = float(sales_df["gross_revenue_allocated_usd"].sum())
    sales_profit = float(sales_df["profit_usd"].sum())
    sales_orders = int(sales_df["order_no"].nunique())
    return_orders = int(df.loc[return_mask, "order_no"].nunique())
    revenue_excluded_orders = int(df.loc[revenue_excluded_mask, "order_no"].nunique())
    after_sales_orders = int(df.loc[after_sales_mask, "order_no"].nunique())
    estimated_orders = int(df.loc[estimated_mask, "order_no"].nunique())
    return {
        "pnlMode": mode,
        "estimateMethod": estimate_method,
        "orders": int(df["order_no"].nunique()),
        "salesOrders": sales_orders,
        "returnOrders": return_orders,
        "revenueExcludedOrders": revenue_excluded_orders,
        "estimatedOrders": estimated_orders,
        "estimatedLines": int(estimated_mask.sum()),
        "estimateMissingOrders": int(df.loc[df.get("pnl_estimate_missing", pd.Series(False, index=df.index)).fillna(False).astype(bool), "order_no"].nunique()),
        "estimateMissingLines": int(df.get("pnl_estimate_missing", pd.Series(False, index=df.index)).fillna(False).astype(bool).sum()),
        "afterSalesOrders": after_sales_orders,
        "lines": int(len(df)),
        "baseRevenue": round(float(df["base_revenue_allocated_usd"].sum()), 2),
        "shippingRevenue": round(float(df["shipping_fee_allocated_usd"].sum()), 2),
        "grossRevenue": round(gross, 2),
        "performanceFee": round(float(df["performance_service_charge_allocated_usd"].sum()), 2),
        "productCost": round(float(df.get("pnl_product_cost_usd", df["product_cost_rule_usd"]).sum()), 2),
        "packagingFee": round(float(df.get("pnl_packaging_fee_usd", df["packaging_fee_rule_usd"]).sum()), 2),
        "afterSalesCost": round(float(df.get("after_sales_cost_usd", pd.Series(0.0, index=df.index)).sum()), 2),
        "profit": round(profit, 2),
        "margin": round(sales_profit / sales_gross, 4) if sales_gross else None,
    }


def add_profit_comparison(summary: dict[str, Any], actual_summary: dict[str, Any], estimated_summary: dict[str, Any]) -> dict[str, Any]:
    actual_profit = float(actual_summary.get("profit") or 0)
    estimated_profit = float(estimated_summary.get("profit") or 0)
    return {
        **summary,
        "actualProfit": round(actual_profit, 2),
        "estimatedProfit": round(estimated_profit, 2),
        "profitDelta": round(estimated_profit - actual_profit, 2),
        "actualMargin": actual_summary.get("margin"),
        "estimatedMargin": estimated_summary.get("margin"),
        "estimatedComparisonOrders": int(estimated_summary.get("estimatedOrders") or 0),
        "estimatedComparisonLines": int(estimated_summary.get("estimatedLines") or 0),
        "estimateMissingComparisonOrders": int(estimated_summary.get("estimateMissingOrders") or 0),
        "estimateMissingComparisonLines": int(estimated_summary.get("estimateMissingLines") or 0),
    }


@app.get("/api/filters")
def api_filters(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=VIEW_PROFIT)
    df = enrich_pnl_items_with_relations(load_items())
    skus = (
        df.groupby(["warehouse_sku_key", "warehouse_sku_label", "warehouse_mapping_status"], dropna=False)
        .agg(
            lines=("goods_id", "count"),
            profit=("profit_usd", "sum"),
            revenue=("gross_revenue_allocated_usd", "sum"),
            sheinSkuCount=("shein_sku_code", nunique_nonempty),
            missingLines=("warehouse_mapping_missing", "sum"),
        )
        .reset_index()
        .sort_values(["profit", "warehouse_sku_label"], ascending=[False, True])
    )
    min_dt = df["order_created_dt"].min()
    max_dt = df["order_created_dt"].max()
    return JSONResponse({
        "shop": load_settings().shop_key,
        "minTime": "" if pd.isna(min_dt) else min_dt.floor("h").strftime("%Y-%m-%dT%H:%M"),
        "maxTime": "" if pd.isna(max_dt) else max_dt.ceil("h").strftime("%Y-%m-%dT%H:%M"),
        "skus": [
            {
                "key": str(row.warehouse_sku_key),
                "label": str(row.warehouse_sku_label),
                "status": str(row.warehouse_mapping_status),
                "lines": int(row.lines),
                "profit": round(float(row.profit or 0), 2),
                "revenue": round(float(row.revenue or 0), 2),
                "sheinSkuCount": int(row.sheinSkuCount or 0),
                "missingLines": int(row.missingLines or 0),
            }
            for row in skus.itertuples(index=False)
        ],
    })


def run_json_script(
    command: list[str],
    *,
    lock: threading.Lock,
    busy_message: str,
    failure_message: str,
    timeout_message: str,
) -> dict[str, Any]:
    if not lock.acquire(blocking=False):
        raise HTTPException(status_code=409, detail=busy_message)
    try:
        try:
            result = subprocess.run(
                command,
                cwd=str(BASE_DIR),
                capture_output=True,
                text=True,
                timeout=300,
                check=False,
            )
        except subprocess.TimeoutExpired as exc:
            raise HTTPException(status_code=504, detail=timeout_message) from exc
    finally:
        lock.release()
    stdout = (result.stdout or "").strip()
    stderr = (result.stderr or "").strip()
    if result.returncode != 0:
        raise HTTPException(
            status_code=500,
            detail={
                "message": failure_message,
                "returnCode": result.returncode,
                "stdout": stdout[-4000:],
                "stderr": stderr[-4000:],
            },
        )
    try:
        return json.loads(stdout) if stdout else {}
    except json.JSONDecodeError:
        return {"stdout": stdout[-4000:]}


def refresh_pnl_export() -> dict[str, Any]:
    script_path = BASE_DIR / "scripts" / "export_orders_profit.py"
    if not script_path.exists():
        raise HTTPException(status_code=500, detail=f"missing export script: {script_path}")
    return run_json_script(
        [sys.executable, str(script_path)],
        lock=REFRESH_EXPORT_LOCK,
        busy_message="profit export is already running",
        failure_message="profit export failed",
        timeout_message="profit export timed out after 300 seconds",
    )


def sync_latest_dataset(data_type: str) -> dict[str, Any]:
    script_path = BASE_DIR / "scripts" / "sync_latest_shein_data.py"
    if not script_path.exists():
        raise HTTPException(status_code=500, detail=f"missing sync script: {script_path}")
    locks = {
        "orders": SYNC_LATEST_ORDERS_LOCK,
        "returns": SYNC_LATEST_RETURNS_LOCK,
    }
    if data_type not in locks:
        raise ValueError(f"unsupported latest sync data type: {data_type}")
    label = "order" if data_type == "orders" else "return"
    return run_json_script(
        [sys.executable, str(script_path), "--data", data_type],
        lock=locks[data_type],
        busy_message=f"latest {label} sync is already running",
        failure_message=f"latest {label} sync failed",
        timeout_message=f"latest {label} sync timed out after 300 seconds",
    )


def sync_order_backed_page_data() -> dict[str, Any]:
    sync = sync_latest_dataset("orders")
    export = refresh_pnl_export()
    return {"status": "ok", "sync": sync, "export": export}


def sync_return_page_data() -> dict[str, Any]:
    sync = sync_latest_dataset("returns")
    export = refresh_pnl_export()
    return {"status": "ok", "sync": sync, "export": export}


@app.post("/api/refresh-pnl-export")
def api_refresh_pnl_export(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=VIEW_PROFIT)
    return JSONResponse({"status": "ok", "export": refresh_pnl_export()})


@app.post("/api/sync-latest-orders")
def api_sync_latest_orders(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=VIEW_PROFIT)
    return JSONResponse({"status": "ok", "sync": sync_latest_dataset("orders")})


@app.post("/api/returns/sync-latest")
def api_sync_latest_returns(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=VIEW_RETURNS)
    return JSONResponse(sync_return_page_data())


@app.post("/api/shipping-fee/sync-latest")
def api_sync_latest_shipping_fee(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=VIEW_SHIPPING_FEE)
    return JSONResponse(sync_order_backed_page_data())


@app.post("/api/logistics/sync-latest")
def api_sync_latest_logistics(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=VIEW_LOGISTICS)
    return JSONResponse(sync_order_backed_page_data())


@app.get("/api/data")
def api_data(token: str | None = None, sku: list[str] | None = Query(default=None), start: str | None = None, end: str | None = None, pnlMode: str | None = None, estimateMethod: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=VIEW_PROFIT)
    mode = normalize_pnl_mode(pnlMode)
    estimate_method = normalize_estimate_method(estimateMethod)
    all_items = enrich_pnl_items_with_relations(load_items())
    sku_fee_averages = build_sku_fulfillment_fee_averages(all_items)
    filtered_items = filter_items(all_items, skus=sku or [], start=start, end=end)
    actual_df = apply_pnl_mode(filtered_items, mode=ACTUAL_PNL_MODE, estimate_method=estimate_method, sku_fee_averages=sku_fee_averages)
    estimated_df = apply_pnl_mode(filtered_items, mode=ESTIMATED_PNL_MODE, estimate_method=estimate_method, sku_fee_averages=sku_fee_averages)
    df = estimated_df if mode == ESTIMATED_PNL_MODE else actual_df
    actual_summary = summarize(actual_df, mode=ACTUAL_PNL_MODE, estimate_method=estimate_method)
    estimated_summary = summarize(estimated_df, mode=ESTIMATED_PNL_MODE, estimate_method=estimate_method)
    summary = add_profit_comparison(summarize(df, mode=mode, estimate_method=estimate_method), actual_summary, estimated_summary)
    summary.update(missing_mapping_summary(df))
    summary["warehouseSkuCount"] = int(df["warehouse_sku_key"].nunique()) if not df.empty else 0
    summary["mappedWarehouseSkuCount"] = int(df.loc[~df["warehouse_mapping_missing"], "warehouse_sku"].nunique()) if not df.empty else 0
    summary["skcCount"] = nunique_nonempty(df["skc_name"]) if not df.empty else 0
    if df.empty:
        return JSONResponse({"summary": summary, "skuTable": [], "skcTable": [], "hourly": [], "orders": []})

    policy = df.get("pnl_policy", pd.Series("standard", index=df.index)).fillna("standard")
    sales_mask = ~policy.isin(SALES_EXCLUDED_POLICIES)
    estimated_mask = df.get("is_estimated_pnl", pd.Series(False, index=df.index)).fillna(False).astype(bool)
    df = df.assign(
        sales_order_no=df["order_no"].where(sales_mask, pd.NA),
        estimated_order_no=df["order_no"].where(estimated_mask, pd.NA),
        sales_line_count=sales_mask.astype(int),
        sales_gross_revenue=df["gross_revenue_allocated_usd"].where(sales_mask, 0.0),
        sales_profit=df["profit_usd"].where(sales_mask, 0.0),
    )

    sku_table = df.groupby(["warehouse_sku_key", "warehouse_sku_label", "warehouse_mapping_status"], dropna=False).agg(
        warehouseSku=("warehouse_sku", first_clean_series),
        supplierSkus=("supplier_sku", compact_series),
        supplierCodes=("supplier_code", compact_series),
        lines=("goods_id", "count"),
        salesLines=("sales_line_count", "sum"),
        estimatedLines=("is_estimated_pnl", "sum"),
        estimateMissingLines=("pnl_estimate_missing", "sum"),
        orders=("order_no", "nunique"),
        salesOrders=("sales_order_no", "nunique"),
        estimatedOrders=("estimated_order_no", "nunique"),
        sheinSkuCount=("shein_sku_code", nunique_nonempty),
        sheinSkus=("shein_sku_code", compact_series),
        skcCount=("skc_name", nunique_nonempty),
        skcs=("skc_name", compact_series),
        spuCount=("spu_name", nunique_nonempty),
        spus=("spu_name", compact_series),
        pcs=("pcs", compact_series),
        title=("product_title", first_clean_series),
        imageUrl=("product_image_url", first_clean_series),
        missingMappingLines=("warehouse_mapping_missing", "sum"),
        weight=("goods_weight", "sum"),
        baseRevenue=("base_revenue_allocated_usd", "sum"),
        shippingRevenue=("shipping_fee_allocated_usd", "sum"),
        grossRevenue=("gross_revenue_allocated_usd", "sum"),
        performanceFee=("performance_service_charge_allocated_usd", "sum"),
        avgFulfillmentFee=("sku_avg_fulfillment_fee_usd", "mean"),
        productCost=("pnl_product_cost_usd", "sum"),
        packagingFee=("pnl_packaging_fee_usd", "sum"),
        afterSalesCost=("after_sales_cost_usd", "sum"),
        profit=("profit_usd", "sum"),
        salesGrossRevenue=("sales_gross_revenue", "sum"),
        salesProfit=("sales_profit", "sum"),
    ).reset_index()
    sku_table["warehouseSkuKey"] = sku_table["warehouse_sku_key"]
    sku_table["warehouseSkuLabel"] = sku_table["warehouse_sku_label"]
    sku_table["mappingStatus"] = sku_table["warehouse_mapping_status"]
    sku_details = (
        df.groupby(["warehouse_sku_key", "warehouse_sku_label", "warehouse_mapping_status"], dropna=False)
        .apply(sku_detail_rows)
        .reset_index(name="sheinSkuDetails")
    )
    sku_table = sku_table.merge(
        sku_details,
        on=["warehouse_sku_key", "warehouse_sku_label", "warehouse_mapping_status"],
        how="left",
    )
    sku_table["margin"] = sku_table["salesProfit"] / sku_table["salesGrossRevenue"].replace({0: pd.NA})
    sku_table["avgProfit"] = sku_table["salesProfit"] / sku_table["salesLines"].replace({0: pd.NA})
    sku_table = sku_table.sort_values("profit", ascending=False)

    skc_table = df.groupby(["skc_key", "skc_label"], dropna=False).agg(
        skcName=("skc_name", first_clean_series),
        spuName=("spu_name", first_clean_series),
        categoryId=("category_id", first_clean_series),
        skcShelfStatus=("skc_shelf_status", "first"),
        title=("product_title", first_clean_series),
        imageUrl=("product_image_url", first_clean_series),
        warehouseSkuCount=("warehouse_sku_key", nunique_nonempty),
        warehouseSkus=("warehouse_sku_label", compact_series),
        sheinSkuCount=("shein_sku_code", nunique_nonempty),
        sheinSkus=("shein_sku_code", compact_series),
        supplierSkus=("supplier_sku", compact_series),
        supplierCodes=("supplier_code", compact_series),
        missingMappingLines=("warehouse_mapping_missing", "sum"),
        lines=("goods_id", "count"),
        salesLines=("sales_line_count", "sum"),
        estimatedLines=("is_estimated_pnl", "sum"),
        estimateMissingLines=("pnl_estimate_missing", "sum"),
        orders=("order_no", "nunique"),
        salesOrders=("sales_order_no", "nunique"),
        estimatedOrders=("estimated_order_no", "nunique"),
        weight=("goods_weight", "sum"),
        baseRevenue=("base_revenue_allocated_usd", "sum"),
        shippingRevenue=("shipping_fee_allocated_usd", "sum"),
        grossRevenue=("gross_revenue_allocated_usd", "sum"),
        performanceFee=("performance_service_charge_allocated_usd", "sum"),
        avgFulfillmentFee=("sku_avg_fulfillment_fee_usd", "mean"),
        productCost=("pnl_product_cost_usd", "sum"),
        packagingFee=("pnl_packaging_fee_usd", "sum"),
        afterSalesCost=("after_sales_cost_usd", "sum"),
        profit=("profit_usd", "sum"),
        salesGrossRevenue=("sales_gross_revenue", "sum"),
        salesProfit=("sales_profit", "sum"),
    ).reset_index()
    skc_table["skcKey"] = skc_table["skc_key"]
    skc_table["skcLabel"] = skc_table["skc_label"]
    skc_table["shelfStatusLabel"] = skc_table["skcShelfStatus"].map({1: "上架", 0: "下架"}).fillna("未知")
    skc_warehouse_details = (
        df.groupby(["skc_key", "skc_label"], dropna=False)
        .apply(skc_warehouse_rows)
        .reset_index(name="warehouseSkuRows")
    )
    skc_table = skc_table.merge(
        skc_warehouse_details,
        on=["skc_key", "skc_label"],
        how="left",
    )
    skc_table["margin"] = skc_table["salesProfit"] / skc_table["salesGrossRevenue"].replace({0: pd.NA})
    skc_table["avgProfit"] = skc_table["salesProfit"] / skc_table["salesLines"].replace({0: pd.NA})
    skc_table = skc_table.sort_values("profit", ascending=False)

    hourly = df.assign(hour=df["order_created_dt"].dt.floor("h")).groupby("hour", dropna=False).agg(
        orders=("order_no", "nunique"),
        estimatedOrders=("estimated_order_no", "nunique"),
        revenue=("gross_revenue_allocated_usd", "sum"),
        profit=("profit_usd", "sum"),
    ).reset_index().sort_values("hour")

    order_table = df.groupby("order_no", dropna=False).agg(
        orderCreated=("order_created_at", "first"),
        orderStatus=("order_status_label", "first"),
        pnlPolicy=("pnl_policy", "first"),
        isEstimated=("is_estimated_pnl", "max"),
        estimatedLines=("is_estimated_pnl", "sum"),
        estimateMissingLines=("pnl_estimate_missing", "sum"),
        estimatedFulfillmentFee=("estimated_fulfillment_fee_allocated_usd", "sum"),
        returnDetected=("return_detected", "max"),
        returnStatus=("return_status", compact_series),
        lines=("goods_id", "count"),
        skus=("warehouse_sku_label", compact_series),
        sheinSkus=("shein_sku_code", compact_series),
        skcs=("skc_name", compact_series),
        missingMappingLines=("warehouse_mapping_missing", "sum"),
        grossRevenue=("gross_revenue_allocated_usd", "sum"),
        performanceFee=("performance_service_charge_allocated_usd", "sum"),
        productCost=("pnl_product_cost_usd", "sum"),
        packagingFee=("pnl_packaging_fee_usd", "sum"),
        afterSalesCost=("after_sales_cost_usd", "sum"),
        profit=("profit_usd", "sum"),
    ).reset_index()
    order_table["margin"] = order_table["profit"] / order_table["grossRevenue"].replace({0: pd.NA})
    order_table = order_table.sort_values("profit", ascending=False).head(2000)

    return JSONResponse({
        "summary": summary,
        "skuTable": [serialize_row(row) for row in sku_table.to_dict(orient="records")],
        "skcTable": [serialize_row(row) for row in skc_table.to_dict(orient="records")],
        "hourly": [serialize_row(row) for row in hourly.to_dict(orient="records")],
        "orders": [serialize_row(row) for row in order_table.to_dict(orient="records")],
    })



def normalize_shipping_fee_dimension(value: str | None) -> str:
    allowed = {item["value"] for item in SHIPPING_FEE_DIMENSIONS}
    value = (value or SHIPPING_FEE_DEFAULT_DIMENSION).strip()
    return value if value in allowed else SHIPPING_FEE_DEFAULT_DIMENSION


def normalize_int_range(value: int | None, *, default: int, min_value: int, max_value: int) -> int:
    if value is None:
        return default
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return max(min_value, min(max_value, parsed))


def shipping_fee_rate(hit_lines: Any, total_lines: Any) -> float | None:
    total = number(total_lines)
    if total <= 0:
        return None
    return number(hit_lines) / total


def none_if_na(value: Any) -> Any:
    try:
        return None if pd.isna(value) else value
    except (TypeError, ValueError):
        return value


def float_or_none(value: Any) -> float | None:
    value = none_if_na(value)
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def int_or_zero(value: Any) -> int:
    value = none_if_na(value)
    if value is None:
        return 0
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return 0


def shipping_fee_effective_items() -> pd.DataFrame:
    df = enrich_pnl_items_with_relations(load_items())
    if df.empty:
        return df
    policy = df.get("pnl_policy", pd.Series("standard", index=df.index)).fillna("standard")
    effective_mask = ~policy.isin(SALES_EXCLUDED_POLICIES)
    df = df.loc[effective_mask & df["order_created_dt"].notna()].copy()
    if df.empty:
        return df
    shipping_source = df.get("shipping_fee_allocated_usd", pd.Series(0.0, index=df.index))
    df["shipping_fee_value_usd"] = pd.to_numeric(shipping_source, errors="coerce").fillna(0.0)
    df["shipping_fee_hit"] = df["shipping_fee_value_usd"].gt(0)
    df["shipping_fee_day"] = df["order_created_dt"].dt.floor("d")
    return df


def dimension_text_series(df: pd.DataFrame, column: str) -> pd.Series:
    if column not in df:
        return pd.Series("", index=df.index, dtype="string")
    return df[column].fillna("").astype(str).str.strip()


def apply_shipping_fee_dimension(df: pd.DataFrame, dimension: str) -> pd.DataFrame:
    dimension = normalize_shipping_fee_dimension(dimension)
    df = df.copy()
    if df.empty:
        df["shipping_fee_group_key"] = pd.Series(dtype="string")
        df["shipping_fee_group_label"] = pd.Series(dtype="string")
        return df

    if dimension == "all":
        df["shipping_fee_group_key"] = "__all__"
        df["shipping_fee_group_label"] = "整体"
        return df

    if dimension == "warehouse_sku":
        key_source = dimension_text_series(df, "warehouse_sku_key")
        label_source = dimension_text_series(df, "warehouse_sku_label")
        missing_key = "__missing_warehouse_sku__"
        missing_label = "未映射仓库 SKU"
    elif dimension == "shein_sku":
        key_source = dimension_text_series(df, "shein_sku_code")
        label_source = key_source.where(key_source.ne(""), dimension_text_series(df, "sku_label"))
        missing_key = "__missing_shein_sku__"
        missing_label = "缺 SHEIN SKU"
    elif dimension == "skc":
        key_source = dimension_text_series(df, "skc_key")
        label_source = dimension_text_series(df, "skc_label")
        missing_key = "__missing_skc__"
        missing_label = "缺商品资料"
    elif dimension == "sku_attr":
        key_source = dimension_text_series(df, "sku_label")
        label_source = key_source
        missing_key = "__missing_sku_attr__"
        missing_label = "未匹配规格"
    elif dimension == "country":
        key_source = dimension_text_series(df, "country")
        label_source = key_source
        missing_key = "__missing_country__"
        missing_label = "未知国家"
    else:
        key_source = dimension_text_series(df, "province")
        label_source = key_source
        missing_key = "__missing_province__"
        missing_label = "未知省州"

    keys: list[str] = []
    labels: list[str] = []
    for key_value, label_value in zip(key_source.tolist(), label_source.tolist()):
        key_text = clean_text(key_value)
        label_text = clean_text(label_value)
        keys.append(key_text or missing_key)
        labels.append(label_text or missing_label)
    df["shipping_fee_group_key"] = keys
    df["shipping_fee_group_label"] = labels
    return df


def shipping_fee_group_metadata(df: pd.DataFrame, dimension: str) -> dict[str, dict[str, str]]:
    if dimension not in {"warehouse_sku", "shein_sku", "skc", "sku_attr"} or df.empty:
        return {}
    metadata: dict[str, dict[str, str]] = {}
    columns = ["shipping_fee_group_key", "product_image_url", "product_title", "shein_sku_code"]
    for group_key, image_url, title, shein_sku in df[columns].itertuples(index=False, name=None):
        key = clean_text(group_key)
        current = metadata.setdefault(key, {"imageUrl": "", "title": "", "sheinSku": ""})
        if not current["imageUrl"]:
            current["imageUrl"] = clean_text(image_url)
        if not current["title"]:
            current["title"] = clean_text(title)
        if not current["sheinSku"]:
            current["sheinSku"] = clean_text(shein_sku)
    return metadata


def shipping_fee_window_summary(
    df: pd.DataFrame,
    *,
    start_day: pd.Timestamp,
    end_day: pd.Timestamp,
    min_periods: int = 1,
) -> dict[str, Any]:
    if df.empty:
        return {"lines": 0, "hitLines": 0, "orders": 0, "shippingRevenue": 0.0, "rate": None}
    window = df.loc[df["shipping_fee_day"].between(start_day, end_day)]
    lines = int(len(window))
    hit_lines = int(window["shipping_fee_hit"].sum()) if lines else 0
    data_start = df["shipping_fee_day"].min()
    available_start = max(start_day, data_start)
    available_periods = max(0, int((end_day - available_start).days) + 1) if end_day >= data_start else 0
    return {
        "lines": lines,
        "hitLines": hit_lines,
        "orders": int(window["order_no"].nunique()) if lines else 0,
        "shippingRevenue": round(float(window["shipping_fee_value_usd"].sum()), 2) if lines else 0.0,
        "rate": shipping_fee_rate(hit_lines, lines) if available_periods >= min_periods else None,
    }


def shipping_fee_date_bounds(df: pd.DataFrame, start: str | None, end: str | None) -> tuple[pd.Timestamp, pd.Timestamp]:
    min_day = df["shipping_fee_day"].min() if not df.empty else pd.Timestamp.now().floor("d")
    max_day = df["shipping_fee_day"].max() if not df.empty else pd.Timestamp.now().floor("d")
    if pd.isna(min_day):
        min_day = pd.Timestamp.now().floor("d")
    if pd.isna(max_day):
        max_day = pd.Timestamp.now().floor("d")
    start_day = parse_datetime(start, min_day).floor("d")
    end_day = parse_datetime(end, max_day).floor("d")
    if end_day < start_day:
        raise HTTPException(status_code=400, detail="end must be >= start")
    return start_day, end_day


def shipping_fee_empty_response(
    *,
    dimension: str,
    rolling_days: int,
    min_periods: int,
    shift_days: int,
    top_n: int,
    start_day: pd.Timestamp | None = None,
    end_day: pd.Timestamp | None = None,
) -> dict[str, Any]:
    start_text = "" if start_day is None else start_day.strftime("%Y-%m-%d")
    end_text = "" if end_day is None else end_day.strftime("%Y-%m-%d")
    return {
        "summary": {
            "dimension": dimension,
            "rollingDays": rolling_days,
            "minPeriods": min_periods,
            "shiftDays": shift_days,
            "topN": top_n,
            "startDate": start_text,
            "endDate": end_text,
            "currentRate": None,
            "shiftedRate": None,
            "changePp": None,
            "currentLines": 0,
            "currentHitLines": 0,
            "shiftedLines": 0,
            "shiftedHitLines": 0,
            "periodLines": 0,
            "periodHitLines": 0,
            "periodRate": None,
            "groupCount": 0,
        },
        "groups": [],
        "series": [],
        "table": [],
    }


@app.get("/api/shipping-fee/filters")
def api_shipping_fee_filters(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=VIEW_SHIPPING_FEE)
    df = shipping_fee_effective_items()
    min_dt = df["shipping_fee_day"].min() if not df.empty else pd.NaT
    max_dt = df["shipping_fee_day"].max() if not df.empty else pd.NaT
    hit_lines = int(df["shipping_fee_hit"].sum()) if not df.empty else 0
    total_lines = int(len(df)) if not df.empty else 0
    return JSONResponse({
        "shop": load_settings().shop_key,
        "minTime": "" if pd.isna(min_dt) else min_dt.strftime("%Y-%m-%dT00:00"),
        "maxTime": "" if pd.isna(max_dt) else max_dt.strftime("%Y-%m-%dT23:59"),
        "dimensions": SHIPPING_FEE_DIMENSIONS,
        "defaults": {
            "dimension": SHIPPING_FEE_DEFAULT_DIMENSION,
            "rollingDays": SHIPPING_FEE_DEFAULT_ROLLING_DAYS,
            "minPeriods": SHIPPING_FEE_DEFAULT_MIN_PERIODS,
            "shiftDays": SHIPPING_FEE_DEFAULT_SHIFT_DAYS,
            "topN": SHIPPING_FEE_DEFAULT_TOP_N,
        },
        "summary": {
            "lines": total_lines,
            "hitLines": hit_lines,
            "rate": shipping_fee_rate(hit_lines, total_lines),
        },
    })


@app.get("/api/shipping-fee/data")
def api_shipping_fee_data(
    token: str | None = None,
    start: str | None = None,
    end: str | None = None,
    dimension: str | None = None,
    rolling: int | None = None,
    minPeriods: int | None = None,
    shift: int | None = None,
    topN: int | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=VIEW_SHIPPING_FEE)
    selected_dimension = normalize_shipping_fee_dimension(dimension)
    rolling_days = normalize_int_range(rolling, default=SHIPPING_FEE_DEFAULT_ROLLING_DAYS, min_value=1, max_value=90)
    min_periods = normalize_int_range(minPeriods, default=rolling_days, min_value=1, max_value=rolling_days)
    shift_days = normalize_int_range(shift, default=SHIPPING_FEE_DEFAULT_SHIFT_DAYS, min_value=1, max_value=90)
    top_n = normalize_int_range(topN, default=SHIPPING_FEE_DEFAULT_TOP_N, min_value=1, max_value=50)

    df = apply_shipping_fee_dimension(shipping_fee_effective_items(), selected_dimension)
    if df.empty:
        return JSONResponse(shipping_fee_empty_response(
            dimension=selected_dimension,
            rolling_days=rolling_days,
            min_periods=min_periods,
            shift_days=shift_days,
            top_n=top_n,
        ))

    start_day, end_day = shipping_fee_date_bounds(df, start, end)
    period_mask = df["shipping_fee_day"].between(start_day, end_day)
    period_df = df.loc[period_mask].copy()
    if period_df.empty:
        return JSONResponse(shipping_fee_empty_response(
            dimension=selected_dimension,
            rolling_days=rolling_days,
            min_periods=min_periods,
            shift_days=shift_days,
            top_n=top_n,
            start_day=start_day,
            end_day=end_day,
        ))

    current_start = end_day - pd.Timedelta(days=rolling_days - 1)
    shifted_end = end_day - pd.Timedelta(days=shift_days)
    shifted_start = shifted_end - pd.Timedelta(days=rolling_days - 1)

    period_grouped = period_df.groupby(["shipping_fee_group_key", "shipping_fee_group_label"], dropna=False).agg(
        periodLines=("shipping_fee_hit", "size"),
        periodHitLines=("shipping_fee_hit", "sum"),
        periodOrders=("order_no", "nunique"),
        periodShippingRevenue=("shipping_fee_value_usd", "sum"),
    ).reset_index()
    period_grouped["periodRate"] = period_grouped["periodHitLines"] / period_grouped["periodLines"].replace({0: pd.NA})
    group_metadata = shipping_fee_group_metadata(period_df, selected_dimension)
    period_grouped = period_grouped.sort_values(
        ["periodOrders", "periodShippingRevenue", "shipping_fee_group_label"],
        ascending=[False, False, True],
    )
    if selected_dimension != "all":
        period_grouped = period_grouped.head(top_n)

    top_groups = [
        {
            "key": clean_text(row.shipping_fee_group_key),
            "label": clean_text(row.shipping_fee_group_label),
            "periodLines": int(row.periodLines or 0),
            "periodHitLines": int(row.periodHitLines or 0),
            "periodRate": float_or_none(row.periodRate),
            "periodOrders": int(row.periodOrders or 0),
            "periodShippingRevenue": round(float(row.periodShippingRevenue or 0), 2),
            **group_metadata.get(clean_text(row.shipping_fee_group_key), {"imageUrl": "", "title": "", "sheinSku": ""}),
        }
        for row in period_grouped.itertuples(index=False)
    ]
    top_keys = {item["key"] for item in top_groups}
    trend_df = df.loc[df["shipping_fee_group_key"].isin(top_keys) & df["shipping_fee_day"].le(end_day)].copy()
    daily_source = trend_df.groupby(
        ["shipping_fee_group_key", "shipping_fee_group_label", "shipping_fee_day"],
        dropna=False,
    ).agg(
        lines=("shipping_fee_hit", "size"),
        hitLines=("shipping_fee_hit", "sum"),
        shippingRevenue=("shipping_fee_value_usd", "sum"),
    ).reset_index()
    daily_source["rate"] = daily_source["hitLines"] / daily_source["lines"].replace({0: pd.NA})

    period_by_key = {item["key"]: item for item in top_groups}
    display_days = pd.date_range(start_day, end_day, freq="D")
    series_rows: list[dict[str, Any]] = []
    table_rows: list[dict[str, Any]] = []
    for group in top_groups:
        group_key = group["key"]
        group_label = group["label"]
        group_daily = daily_source.loc[daily_source["shipping_fee_group_key"].eq(group_key)].sort_values("shipping_fee_day")
        frame = pd.DataFrame({"day": display_days})
        in_range = group_daily.loc[group_daily["shipping_fee_day"].between(start_day, end_day)]
        frame = frame.merge(in_range, left_on="day", right_on="shipping_fee_day", how="left")
        prior_rates = group_daily.loc[group_daily["shipping_fee_day"].lt(start_day), "rate"].dropna()
        last_rate = float_or_none(prior_rates.iloc[-1]) if not prior_rates.empty else None
        filled_rates: list[float | None] = []
        for value in frame["rate"].tolist():
            current_value = float_or_none(value)
            if current_value is not None:
                last_rate = current_value
            filled_rates.append(last_rate)
        frame["rate"] = filled_rates
        frame["shippingRevenue"] = pd.to_numeric(frame["shippingRevenue"], errors="coerce").fillna(0.0)
        for row in frame.to_dict(orient="records"):
            series_rows.append({
                "day": row["day"],
                "groupKey": group_key,
                "groupLabel": group_label,
                "rate": float_or_none(row.get("rate")),
                "shippingRevenue": round(number(row.get("shippingRevenue")), 2),
            })

        group_df = df.loc[df["shipping_fee_group_key"].eq(group_key)]
        current = shipping_fee_window_summary(
            group_df,
            start_day=current_start,
            end_day=end_day,
            min_periods=min_periods,
        )
        shifted = shipping_fee_window_summary(
            group_df,
            start_day=shifted_start,
            end_day=shifted_end,
            min_periods=min_periods,
        )
        period = period_by_key.get(group_key, {})
        current_rate = current["rate"]
        shifted_rate = shifted["rate"]
        table_rows.append({
            "groupKey": group_key,
            "groupLabel": group_label,
            "currentRate": current_rate,
            "shiftedRate": shifted_rate,
            "changePp": None if current_rate is None or shifted_rate is None else current_rate - shifted_rate,
            "periodRate": period.get("periodRate"),
            "periodOrders": int(period.get("periodOrders") or 0),
            "periodShippingRevenue": period.get("periodShippingRevenue") or 0.0,
        })

    table_rows = sorted(
        table_rows,
        key=lambda item: (item["periodOrders"], number(item["periodShippingRevenue"]), clean_text(item["groupLabel"])),
        reverse=True,
    )
    current_summary = shipping_fee_window_summary(
        df, start_day=current_start, end_day=end_day, min_periods=min_periods
    )
    shifted_summary = shipping_fee_window_summary(
        df, start_day=shifted_start, end_day=shifted_end, min_periods=min_periods
    )
    period_summary = shipping_fee_window_summary(df, start_day=start_day, end_day=end_day)
    current_rate = current_summary["rate"]
    shifted_rate = shifted_summary["rate"]

    return JSONResponse({
        "summary": {
            "dimension": selected_dimension,
            "rollingDays": rolling_days,
            "minPeriods": min_periods,
            "shiftDays": shift_days,
            "topN": top_n,
            "startDate": start_day.strftime("%Y-%m-%d"),
            "endDate": end_day.strftime("%Y-%m-%d"),
            "currentWindowStart": current_start.strftime("%Y-%m-%d"),
            "currentWindowEnd": end_day.strftime("%Y-%m-%d"),
            "shiftedWindowStart": shifted_start.strftime("%Y-%m-%d"),
            "shiftedWindowEnd": shifted_end.strftime("%Y-%m-%d"),
            "currentRate": current_rate,
            "shiftedRate": shifted_rate,
            "changePp": None if current_rate is None or shifted_rate is None else current_rate - shifted_rate,
            "currentLines": current_summary["lines"],
            "currentHitLines": current_summary["hitLines"],
            "currentOrders": current_summary["orders"],
            "currentShippingRevenue": current_summary["shippingRevenue"],
            "shiftedLines": shifted_summary["lines"],
            "shiftedHitLines": shifted_summary["hitLines"],
            "shiftedOrders": shifted_summary["orders"],
            "shiftedShippingRevenue": shifted_summary["shippingRevenue"],
            "periodLines": period_summary["lines"],
            "periodHitLines": period_summary["hitLines"],
            "periodOrders": period_summary["orders"],
            "periodShippingRevenue": period_summary["shippingRevenue"],
            "periodRate": period_summary["rate"],
            "groupCount": int(period_df["shipping_fee_group_key"].nunique()),
        },
        "groups": [serialize_row(item) for item in top_groups],
        "series": [serialize_row(item) for item in series_rows],
        "table": [serialize_row(item) for item in table_rows],
    })


@app.get("/api/logistics/filters")
def api_logistics_filters(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=VIEW_LOGISTICS)
    df = prepare_logistics_items(load_items())
    effective = df.loc[df["logistics_effective_line"]] if not df.empty else df
    sku_source = effective if not effective.empty else df
    skus: list[dict[str, Any]] = []
    if not sku_source.empty:
        grouped = (
            sku_source.groupby(["logistics_sku_key", "logistics_sku_code", "logistics_sku_label"], dropna=False)
            .agg(
                lines=("goods_id", "count"),
                orders=("order_no", "nunique"),
                fulfillmentFee=("logistics_fulfillment_fee_usd", "sum"),
            )
            .reset_index()
            .sort_values(["fulfillmentFee", "lines", "logistics_sku_label"], ascending=[False, False, True])
        )
        skus = [
            {
                "key": str(row.logistics_sku_key),
                "code": str(row.logistics_sku_code or ""),
                "label": str(row.logistics_sku_label or "未匹配"),
                "lines": int(row.lines),
                "orders": int(row.orders),
                "fulfillmentFee": round(float(row.fulfillmentFee or 0), 2),
            }
            for row in grouped.itertuples(index=False)
        ]
    min_dt = df["order_created_dt"].min() if not df.empty else pd.NaT
    max_dt = df["order_created_dt"].max() if not df.empty else pd.NaT
    return JSONResponse({
        "shop": "default · shein-us",
        "minTime": "" if pd.isna(min_dt) else min_dt.floor("h").strftime("%Y-%m-%dT%H:%M"),
        "maxTime": "" if pd.isna(max_dt) else max_dt.ceil("h").strftime("%Y-%m-%dT%H:%M"),
        "skus": skus,
        "excludedStatuses": sorted(LOGISTICS_EXCLUDED_STATUS_CODES),
    })


@app.get("/api/logistics/data")
def api_logistics_data(
    token: str | None = None,
    sku: list[str] | None = Query(default=None),
    start: str | None = None,
    end: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=VIEW_LOGISTICS)
    all_items = prepare_logistics_items(load_items())
    filtered = filter_logistics_items(all_items, sku_keys=sku or [], start=start, end=end)
    effective = filtered.loc[filtered["logistics_effective_line"]].copy() if not filtered.empty else filtered.copy()
    excluded = filtered.loc[filtered["logistics_excluded_line"]].copy() if not filtered.empty else filtered.copy()

    effective_lines = int(len(effective))
    effective_orders = int(effective["order_no"].nunique()) if not effective.empty else 0
    total_fee = float(effective["logistics_fulfillment_fee_usd"].sum()) if not effective.empty else 0.0
    total_revenue = float(effective["logistics_gross_revenue_usd"].sum()) if not effective.empty else 0.0
    after_sales_mask = effective["logistics_after_sales_line"] if not effective.empty else pd.Series(dtype=bool)
    zero_fee_mask = effective["logistics_zero_fee_line"] if not effective.empty else pd.Series(dtype=bool)
    summary = {
        "totalLines": int(len(filtered)),
        "totalOrders": int(filtered["order_no"].nunique()) if not filtered.empty else 0,
        "effectiveLines": effective_lines,
        "effectiveOrders": effective_orders,
        "excludedLines": int(len(excluded)),
        "excludedOrders": int(excluded["order_no"].nunique()) if not excluded.empty else 0,
        "skuCount": int(effective["logistics_sku_key"].nunique()) if not effective.empty else 0,
        "fulfillmentFee": round(total_fee, 2),
        "avgFeePerLine": round(total_fee / effective_lines, 4) if effective_lines else 0,
        "avgFeePerOrder": round(total_fee / effective_orders, 4) if effective_orders else 0,
        "grossRevenue": round(total_revenue, 2),
        "feeRate": round(total_fee / total_revenue, 4) if total_revenue else None,
        "afterSalesLines": int(after_sales_mask.sum()) if effective_lines else 0,
        "afterSalesOrders": int(effective.loc[after_sales_mask, "order_no"].nunique()) if effective_lines else 0,
        "zeroFeeLines": int(zero_fee_mask.sum()) if effective_lines else 0,
        "zeroFeeOrders": int(effective.loc[zero_fee_mask, "order_no"].nunique()) if effective_lines else 0,
    }

    if effective.empty:
        return JSONResponse({"summary": summary, "skuTable": [], "daily": [], "statusMix": [], "feeBuckets": []})

    group_cols = ["logistics_sku_key", "logistics_sku_code", "logistics_sku_label"]
    sku_table = effective.groupby(group_cols, dropna=False).agg(
        lines=("goods_id", "count"),
        orders=("order_no", "nunique"),
        afterSalesLines=("logistics_after_sales_line", "sum"),
        zeroFeeLines=("logistics_zero_fee_line", "sum"),
        fulfillmentFee=("logistics_fulfillment_fee_usd", "sum"),
        avgFulfillmentFee=("logistics_fulfillment_fee_usd", "mean"),
        medianFulfillmentFee=("logistics_fulfillment_fee_usd", "median"),
        maxFulfillmentFee=("logistics_fulfillment_fee_usd", "max"),
        revenue=("logistics_gross_revenue_usd", "sum"),
        weight=("logistics_weight", "sum"),
    ).reset_index()
    p75 = effective.groupby(group_cols, dropna=False)["logistics_fulfillment_fee_usd"].quantile(0.75).reset_index(name="p75FulfillmentFee")
    sku_table = sku_table.merge(p75, on=group_cols, how="left")
    sku_table["avgFeePerOrder"] = sku_table["fulfillmentFee"] / sku_table["orders"].replace({0: pd.NA})
    sku_table["feeShare"] = sku_table["fulfillmentFee"] / total_fee if total_fee else 0
    sku_table["feeRate"] = sku_table["fulfillmentFee"] / sku_table["revenue"].replace({0: pd.NA})
    sku_table["avgWeight"] = sku_table["weight"] / sku_table["lines"].replace({0: pd.NA})
    sku_table = sku_table.sort_values(["fulfillmentFee", "avgFulfillmentFee"], ascending=[False, False])

    daily = effective.assign(day=effective["order_created_dt"].dt.floor("d")).groupby("day", dropna=False).agg(
        lines=("goods_id", "count"),
        orders=("order_no", "nunique"),
        fulfillmentFee=("logistics_fulfillment_fee_usd", "sum"),
        revenue=("logistics_gross_revenue_usd", "sum"),
    ).reset_index().sort_values("day")
    daily["avgFee"] = daily["fulfillmentFee"] / daily["lines"].replace({0: pd.NA})
    daily["feeRate"] = daily["fulfillmentFee"] / daily["revenue"].replace({0: pd.NA})

    status_frame = filtered.copy()
    status_frame["statusLabel"] = status_frame.get("order_status_label", pd.Series("", index=status_frame.index)).fillna("").astype(str).replace({"": "未知状态"})
    status_mix = status_frame.groupby(["order_status", "statusLabel"], dropna=False).agg(
        lines=("goods_id", "count"),
        orders=("order_no", "nunique"),
        effectiveLines=("logistics_effective_line", "sum"),
        excludedLines=("logistics_excluded_line", "sum"),
        afterSalesLines=("logistics_after_sales_line", "sum"),
        fulfillmentFee=("logistics_fulfillment_fee_usd", "sum"),
    ).reset_index().sort_values(["lines", "statusLabel"], ascending=[False, True])

    return JSONResponse({
        "summary": summary,
        "skuTable": [serialize_row(row) for row in sku_table.to_dict(orient="records")],
        "daily": [serialize_row(row) for row in daily.to_dict(orient="records")],
        "statusMix": [serialize_row(row) for row in status_mix.to_dict(orient="records")],
        "feeBuckets": logistics_fee_buckets(effective["logistics_fulfillment_fee_usd"]),
    })



@app.get("/api/returns/filters")
def api_return_filters(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=VIEW_RETURNS)
    df = load_return_orders()
    min_time, max_time = return_time_bounds(df)
    statuses = []
    if not df.empty:
        grouped = (
            df.assign(status_label=df["return_status_label"].fillna("未知状态"))
            .groupby(["return_status", "status_label"], dropna=False)
            .size()
            .reset_index(name="count")
            .sort_values(["return_status", "status_label"])
        )
        statuses = [
            {"code": clean_text(row.return_status), "label": clean_text(row.status_label) or "未知状态", "count": int(row.count)}
            for row in grouped.itertuples(index=False)
        ]
    return JSONResponse({
        "shop": load_settings().shop_key,
        "minTime": min_time,
        "maxTime": max_time,
        "statuses": statuses,
        "timeFields": [{"value": key, "label": label} for key, (_, label) in RETURN_TIME_FIELDS.items()],
    })


@app.get("/api/returns/data")
def api_return_data(
    token: str | None = None,
    start: str | None = None,
    end: str | None = None,
    status: list[str] | None = Query(default=None),
    timeField: str = "request",
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=VIEW_RETURNS)
    df, time_label = filter_return_orders(load_return_orders(), start=start, end=end, statuses=status or [], time_field=timeField)
    pnl_by_order = order_pnl_lookup()
    rows: list[dict[str, Any]] = []
    if not df.empty:
        df = df.sort_values("return_filter_dt", ascending=False, na_position="last").head(RETURN_PAGE_LIMIT)
        for record in df.to_dict(orient="records"):
            goods = summarize_return_goods(record.get("return_goods_info_list"))
            order_no = clean_text(record.get("order_no"))
            pnl = pnl_by_order.get(order_no, {})
            rows.append({
                "returnNo": clean_text(record.get("return_no")),
                "orderNo": order_no,
                "returnStatus": clean_text(record.get("return_status")),
                "returnStatusLabel": clean_text(record.get("return_status_label")),
                "orderStatusLabel": clean_text(record.get("order_status_label")),
                "orderTypeLabel": clean_text(record.get("order_type_label")),
                "site": clean_text(record.get("site")),
                "requestReturnTime": clean_text(record.get("request_return_time")),
                "lastUpdateTime": clean_text(record.get("last_update_time")),
                "completedTime": clean_text(record.get("completed_time")),
                "sellerSignedTime": clean_text(record.get("seller_signed_time")),
                "cancelTime": clean_text(record.get("cancel_time")),
                "noReturnGoodsSign": clean_text(record.get("no_return_goods_sign")),
                "noReturnGoodsLabel": "不退货" if clean_text(record.get("no_return_goods_sign")) == "1" else "退货",
                "expressCompanyName": clean_text(record.get("express_company_name")),
                "platformExpressNo": clean_text(record.get("platform_express_no")),
                "memberExpressNo": clean_text(record.get("member_express_no")),
                "performanceCost": clean_number(record.get("performance_cost")),
                "goodsLines": goods["goodsLines"],
                "skus": goods["skus"],
                "goodsTitles": goods["goodsTitles"],
                "reasons": goods["reasons"],
                "currency": goods["currency"],
                "sellerAmount": goods["sellerAmount"],
                "goodsPerformancePrice": goods["goodsPerformancePrice"],
                "returnExpense": goods["returnExpense"],
                "estimateIncome": goods["estimateIncome"],
                "pnlInternalCost": pnl.get("pnlInternalCost", 0.0),
                "pnlProfit": pnl.get("pnlProfit", 0.0),
                "originalGrossRevenue": pnl.get("originalGrossRevenue", 0.0),
                "originalProfit": pnl.get("originalProfit", 0.0),
            })
    summary = {
        "returnOrders": len(rows),
        "linkedOrders": len({row["orderNo"] for row in rows if row["orderNo"]}),
        "goodsLines": sum(int(row["goodsLines"] or 0) for row in rows),
        "performanceCost": round(sum(float(row["performanceCost"] or 0) for row in rows), 2),
        "pnlInternalCost": round(sum(float(row["pnlInternalCost"] or 0) for row in rows), 2),
        "pnlProfit": round(sum(float(row["pnlProfit"] or 0) for row in rows), 2),
        "originalGrossRevenue": round(sum(float(row["originalGrossRevenue"] or 0) for row in rows), 2),
        "originalProfit": round(sum(float(row["originalProfit"] or 0) for row in rows), 2),
        "timeFieldLabel": time_label,
        "limit": RETURN_PAGE_LIMIT,
    }
    return JSONResponse({"summary": summary, "rows": rows})


SKU_MAPPING_WAREHOUSE_ALIASES = {
    "warehouse_sku", "warehousesku", "warehouse sku", "仓库sku", "仓库SKU", "仓库 SKU",
    "仓库编码", "仓库商品编码", "本地sku", "本地SKU", "库存sku", "库存SKU", "sku",
}
SKU_MAPPING_SHEIN_ALIASES = {
    "shein_sku", "sheinsku", "shein sku", "SHEIN SKU", "sheinSKU", "平台sku", "平台SKU",
    "shein平台sku", "shein平台SKU", "shein商品sku", "shein商品SKU", "店铺sku", "店铺SKU",
}
SKU_MAPPING_LENGTH_ALIASES = {"length_cm", "lengthcm", "length", "长", "长度", "长cm", "长度cm", "product_length", "package_length"}
SKU_MAPPING_WIDTH_ALIASES = {"width_cm", "widthcm", "width", "宽", "宽度", "宽cm", "宽度cm", "product_width", "package_width"}
SKU_MAPPING_HEIGHT_ALIASES = {"height_cm", "heightcm", "height", "高", "高度", "高cm", "高度cm", "product_height", "package_height"}
SKU_MAPPING_WEIGHT_ALIASES = {"weight_kg", "weightkg", "weight", "重量", "重", "重量kg", "product_weight", "package_weight"}
SKU_MAPPING_PURCHASE_PRICE_ALIASES = {"purchase_price", "purchaseprice", "purchase", "采购价", "采购价格", "买入价", "进货价"}
SKU_MAPPING_OCEAN_FREIGHT_PRICE_ALIASES = {"ocean_freight_price", "oceanfreightprice", "ocean_freight", "sea_freight", "shipping_price", "海运价格", "海运价", "海运费"}
SKU_MAPPING_OPERATION_FEE_ALIASES = {"operation_fee_price", "operationfeeprice", "operation_fee", "operationfee", "handling_fee", "操作费", "操作费用"}
DEFAULT_SKU_GROUP = "default"
SKU_MAPPING_IMPORT_HEADERS = ["warehouse_sku", "shein_sku", "length_cm", "width_cm", "height_cm", "weight_kg", "purchase_price", "ocean_freight_price", "operation_fee_price"]
SKU_MAPPING_HEADER_ALIASES = (
    SKU_MAPPING_WAREHOUSE_ALIASES
    | SKU_MAPPING_SHEIN_ALIASES
    | SKU_MAPPING_LENGTH_ALIASES
    | SKU_MAPPING_WIDTH_ALIASES
    | SKU_MAPPING_HEIGHT_ALIASES
    | SKU_MAPPING_WEIGHT_ALIASES
    | SKU_MAPPING_PURCHASE_PRICE_ALIASES
    | SKU_MAPPING_OCEAN_FREIGHT_PRICE_ALIASES
    | SKU_MAPPING_OPERATION_FEE_ALIASES
)


def ensure_sku_mapping_store() -> tuple[str, str]:
    from .db import ensure_sku_mapping_schema

    url = database_url()
    ensure_sku_mapping_schema(url)
    return url, load_settings().shop_key


def normalize_sku_mapping_header(value: Any) -> str:
    return re.sub(r"[\s_\-()（）/]+", "", str(value or "").strip().lower())


def read_table_value(row: dict[str, Any], aliases: set[str]) -> Any:
    normalized_aliases = {normalize_sku_mapping_header(alias) for alias in aliases}
    for key, value in row.items():
        if normalize_sku_mapping_header(key) in normalized_aliases:
            return value
    return ""


def decode_csv_bytes(content: bytes) -> str:
    for encoding in ("utf-8-sig", "gb18030", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise HTTPException(status_code=400, detail="CSV encoding is not supported")


def table_has_sku_mapping_header(values: list[str]) -> bool:
    aliases = {normalize_sku_mapping_header(value) for value in SKU_MAPPING_HEADER_ALIASES}
    return any(normalize_sku_mapping_header(value) in aliases for value in values)


def sku_mapping_record_from_header_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "warehouse_sku": read_table_value(row, SKU_MAPPING_WAREHOUSE_ALIASES),
        "shein_sku": read_table_value(row, SKU_MAPPING_SHEIN_ALIASES),
        "length_cm": read_table_value(row, SKU_MAPPING_LENGTH_ALIASES),
        "width_cm": read_table_value(row, SKU_MAPPING_WIDTH_ALIASES),
        "height_cm": read_table_value(row, SKU_MAPPING_HEIGHT_ALIASES),
        "weight_kg": read_table_value(row, SKU_MAPPING_WEIGHT_ALIASES),
        "purchase_price": read_table_value(row, SKU_MAPPING_PURCHASE_PRICE_ALIASES),
        "ocean_freight_price": read_table_value(row, SKU_MAPPING_OCEAN_FREIGHT_PRICE_ALIASES),
        "operation_fee_price": read_table_value(row, SKU_MAPPING_OPERATION_FEE_ALIASES),
    }


def parse_sku_mapping_table_rows(rows: list[tuple[int, list[Any]]]) -> list[tuple[int, dict[str, Any]]]:
    nonempty_rows = [
        (row_number, ["" if value is None else value for value in values])
        for row_number, values in rows
        if any(str(value or "").strip() for value in values)
    ]
    if not nonempty_rows:
        return []

    _, first_values = nonempty_rows[0]
    if table_has_sku_mapping_header([str(value or "") for value in first_values]):
        headers = [str(value or "").strip() for value in first_values]
        records: list[tuple[int, dict[str, Any]]] = []
        for row_number, values in nonempty_rows[1:]:
            row = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
            records.append((row_number, sku_mapping_record_from_header_row(row)))
        return records

    return [
        (
            row_number,
            {field: values[index] if index < len(values) else "" for index, field in enumerate(SKU_MAPPING_IMPORT_HEADERS)},
        )
        for row_number, values in nonempty_rows
    ]


def parse_sku_mapping_csv(text: str) -> list[tuple[int, dict[str, Any]]]:
    return parse_sku_mapping_table_rows([
        (row_number, values)
        for row_number, values in enumerate(csv.reader(io.StringIO(text)), start=1)
    ])


def parse_sku_mapping_excel(content: bytes) -> list[tuple[int, dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        raise HTTPException(status_code=500, detail="Excel import requires openpyxl. Run: pip install -r requirements.txt")

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Excel file could not be read: {exc}")
    worksheet = workbook.active
    rows = [
        (row_number, list(values))
        for row_number, values in enumerate(worksheet.iter_rows(values_only=True), start=1)
    ]
    workbook.close()
    return parse_sku_mapping_table_rows(rows)


def is_excel_upload(filename: str, content_type: str) -> bool:
    filename = filename.lower()
    return filename.endswith((".xlsx", ".xlsm")) or content_type in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroenabled.12",
    }


def sku_mapping_text(value: Any, field_name: str) -> str:
    text = str(value or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail=f"{field_name} is required")
    if len(text) > 240:
        raise HTTPException(status_code=400, detail=f"{field_name} is too long")
    return text


def sku_mapping_decimal(value: Any, field_name: str) -> Decimal | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null"}:
        return None
    try:
        decimal_value = Decimal(text)
    except (InvalidOperation, ValueError):
        raise HTTPException(status_code=400, detail=f"{field_name} must be a number")
    if not decimal_value.is_finite():
        raise HTTPException(status_code=400, detail=f"{field_name} must be a finite number")
    if decimal_value < 0:
        raise HTTPException(status_code=400, detail=f"{field_name} cannot be negative")
    return decimal_value


def sku_mapping_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return float(value)


def payload_value(payload: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in payload:
            return payload[key]
    return None


WAREHOUSE_DETAIL_PAYLOAD_KEYS = {
    "length_cm", "lengthCm", "length", "productLength",
    "width_cm", "widthCm", "width", "productWidth",
    "height_cm", "heightCm", "height", "productHeight",
    "weight_kg", "weightKg", "weight", "productWeight",
    "purchase_price", "purchasePrice", "purchase", "purchaseCost",
    "ocean_freight_price", "oceanFreightPrice", "oceanFreight", "seaFreight", "shippingPrice",
    "operation_fee_price", "operationFeePrice", "operationFee", "handlingFee",
}


def warehouse_detail_payload_present(payload: dict[str, Any]) -> bool:
    return any(key in payload for key in WAREHOUSE_DETAIL_PAYLOAD_KEYS)


def normalize_warehouse_sku_payload(payload: dict[str, Any]) -> dict[str, Any]:
    return {
        "warehouse_sku": sku_mapping_text(payload_value(payload, "warehouse_sku", "warehouseSku"), "warehouse_sku"),
        "length_cm": sku_mapping_decimal(payload_value(payload, "length_cm", "lengthCm", "length", "productLength"), "length_cm"),
        "width_cm": sku_mapping_decimal(payload_value(payload, "width_cm", "widthCm", "width", "productWidth"), "width_cm"),
        "height_cm": sku_mapping_decimal(payload_value(payload, "height_cm", "heightCm", "height", "productHeight"), "height_cm"),
        "weight_kg": sku_mapping_decimal(payload_value(payload, "weight_kg", "weightKg", "weight", "productWeight"), "weight_kg"),
        "cost_price": None,
        "purchase_price": sku_mapping_decimal(payload_value(payload, "purchase_price", "purchasePrice", "purchase", "purchaseCost"), "purchase_price"),
        "ocean_freight_price": sku_mapping_decimal(payload_value(payload, "ocean_freight_price", "oceanFreightPrice", "oceanFreight", "seaFreight", "shippingPrice"), "ocean_freight_price"),
        "operation_fee_price": sku_mapping_decimal(payload_value(payload, "operation_fee_price", "operationFeePrice", "operationFee", "handlingFee"), "operation_fee_price"),
        "note": clean_text(payload_value(payload, "note")) or None,
        "enabled": bool(payload_value(payload, "enabled")) if payload_value(payload, "enabled") is not None else True,
    }


def normalize_sku_mapping_payload(payload: dict[str, Any]) -> dict[str, Any]:
    return {
        "warehouse_sku": sku_mapping_text(payload_value(payload, "warehouse_sku", "warehouseSku"), "warehouse_sku"),
        "shein_sku": sku_mapping_text(payload_value(payload, "shein_sku", "sheinSku"), "shein_sku"),
        "sku_group": DEFAULT_SKU_GROUP,
        "warehouse_qty": Decimal("1"),
        "length_cm": None,
        "width_cm": None,
        "height_cm": None,
        "weight_kg": None,
        "cost_price": None,
        "purchase_price": None,
        "ocean_freight_price": None,
        "note": None,
        "enabled": True,
    }


def sku_meta_value(sku_meta: dict[str, Any] | None, shein_sku: str) -> dict[str, Any]:
    value = (sku_meta or {}).get(shein_sku, {})
    if isinstance(value, str):
        return {"label": value}
    return value if isinstance(value, dict) else {}


def serialize_sku_mapping(row: dict[str, Any], sku_meta: dict[str, Any] | None = None) -> dict[str, Any]:
    updated_at = row.get("updated_at")
    shein_sku = clean_text(row.get("shein_sku"))
    meta = sku_meta_value(sku_meta, shein_sku)
    sku_label = clean_text(meta.get("label"))
    return {
        "id": row.get("id"),
        "warehouseSku": clean_text(row.get("warehouse_sku")),
        "sheinSku": shein_sku,
        "sheinSkuLabel": sku_label,
        "imageUrl": clean_text(meta.get("imageUrl")),
        "title": clean_text(meta.get("title")),
        "skcName": clean_text(meta.get("skcName")),
        "spuName": clean_text(meta.get("spuName")),
        "lengthCm": sku_mapping_number(row.get("length_cm")),
        "widthCm": sku_mapping_number(row.get("width_cm")),
        "heightCm": sku_mapping_number(row.get("height_cm")),
        "weightKg": sku_mapping_number(row.get("weight_kg")),
        "purchasePrice": sku_mapping_number(row.get("purchase_price")),
        "oceanFreightPrice": sku_mapping_number(row.get("ocean_freight_price")),
        "updatedAt": updated_at.strftime("%Y-%m-%d %H:%M") if hasattr(updated_at, "strftime") else clean_text(updated_at),
    }


def summarize_sku_mappings(rows: list[dict[str, Any]]) -> dict[str, Any]:
    warehouse_skus = {clean_text(row.get("warehouse_sku")) for row in rows if clean_text(row.get("warehouse_sku"))}
    shein_skus = {clean_text(row.get("shein_sku")) for row in rows if clean_text(row.get("shein_sku"))}
    return {"total": len(rows), "warehouseSkus": len(warehouse_skus), "sheinSkus": len(shein_skus)}


def serialize_warehouse_sku(row: dict[str, Any]) -> dict[str, Any]:
    updated_at = row.get("updated_at")
    return {
        "id": row.get("id"),
        "warehouseSku": clean_text(row.get("warehouse_sku")),
        "lengthCm": sku_mapping_number(row.get("length_cm")),
        "widthCm": sku_mapping_number(row.get("width_cm")),
        "heightCm": sku_mapping_number(row.get("height_cm")),
        "weightKg": sku_mapping_number(row.get("weight_kg")),
        "purchasePrice": sku_mapping_number(row.get("purchase_price")),
        "oceanFreightPrice": sku_mapping_number(row.get("ocean_freight_price")),
        "operationFeePrice": sku_mapping_number(row.get("operation_fee_price")) or 0.0,
        "note": clean_text(row.get("note")),
        "enabled": bool(row.get("enabled", True)),
        "updatedAt": updated_at.strftime("%Y-%m-%d %H:%M") if hasattr(updated_at, "strftime") else clean_text(updated_at),
    }


def default_warehouse_sku_response(settings: dict[str, Any]) -> dict[str, Any]:
    updated_at = settings.get("updated_at")
    return {
        "warehouseSku": clean_text(settings.get("default_warehouse_sku")),
        "updatedAt": updated_at.strftime("%Y-%m-%d %H:%M") if hasattr(updated_at, "strftime") else clean_text(updated_at),
    }


def unmapped_shein_sku_rows(shein_skus: list[dict[str, Any]], rows: list[dict[str, Any]], default_warehouse_sku: str) -> list[dict[str, Any]]:
    mapped = {clean_text(row.get("shein_sku")) for row in rows if clean_text(row.get("shein_sku"))}
    return [
        {
            "sheinSku": item["sku"],
            "sheinSkuLabel": item["label"],
            "warehouseSku": default_warehouse_sku,
            "imageUrl": clean_text(item.get("imageUrl")),
            "title": clean_text(item.get("title")),
            "skcName": clean_text(item.get("skcName")),
            "spuName": clean_text(item.get("spuName")),
            "lines": item["lines"],
            "orders": item["orders"],
        }
        for item in shein_skus
        if clean_text(item.get("sku")) and clean_text(item.get("sku")) not in mapped
    ]


def product_sku_label(row: dict[str, Any]) -> str:
    title = clean_text(row.get("title"))
    supplier_sku = clean_text(row.get("supplier_sku"))
    supplier_code = clean_text(row.get("supplier_code"))
    skc_name = clean_text(row.get("skc_name"))
    spu_name = clean_text(row.get("spu_name"))
    parts = [part for part in (title, supplier_sku, supplier_code, skc_name, spu_name) if part]
    return " · ".join(parts[:3]) if parts else clean_text(row.get("sku_code"))


def product_sku_options() -> list[dict[str, Any]]:
    try:
        from .db import list_product_sku_options

        rows = list_product_sku_options(database_url(), shop_key=load_settings().shop_key)
    except Exception:
        return []
    return [
        {
            "sku": clean_text(row.get("sku_code")),
            "label": product_sku_label(row),
            "imageUrl": clean_text(row.get("main_pic_url")),
            "title": clean_text(row.get("title")),
            "skcName": clean_text(row.get("skc_name")),
            "spuName": clean_text(row.get("spu_name")),
            "supplierSku": clean_text(row.get("supplier_sku")),
            "supplierCode": clean_text(row.get("supplier_code")),
            "lines": 0,
            "orders": 0,
            "source": "product",
        }
        for row in rows
        if clean_text(row.get("sku_code"))
    ]


def shein_sku_options() -> list[dict[str, Any]]:
    options: dict[str, dict[str, Any]] = {}
    if ITEMS_PATH.exists():
        df = load_items()
        if not df.empty:
            sku_col = "sku_code" if "sku_code" in df.columns else "sku_label"
            label_col = "sku_label" if "sku_label" in df.columns else "sku_attr_us"
            grouped = (
                df.assign(
                    shein_sku=df[sku_col].fillna("").astype(str).str.strip(),
                    label=df[label_col].fillna("").astype(str).str.strip(),
                )
                .loc[lambda x: x["shein_sku"].ne("")]
                .groupby(["shein_sku", "label"], dropna=False)
                .agg(lines=("goods_id", "count"), orders=("order_no", "nunique"))
                .reset_index()
                .sort_values(["lines", "shein_sku"], ascending=[False, True])
            )
            for row in grouped.itertuples(index=False):
                sku = str(row.shein_sku)
                options[sku] = {
                    "sku": sku,
                    "label": str(row.label or row.shein_sku),
                    "imageUrl": "",
                    "title": "",
                    "skcName": "",
                    "spuName": "",
                    "supplierSku": "",
                    "supplierCode": "",
                    "lines": int(row.lines),
                    "orders": int(row.orders),
                    "source": "orders",
                }
    for item in product_sku_options():
        sku = item["sku"]
        if sku in options:
            options[sku]["source"] = "orders+product"
            if options[sku]["label"] == sku and item.get("label"):
                options[sku]["label"] = item["label"]
            for key in ("imageUrl", "title", "skcName", "spuName", "supplierSku", "supplierCode"):
                if item.get(key) and not options[sku].get(key):
                    options[sku][key] = item[key]
        else:
            options[sku] = item
    return sorted(options.values(), key=lambda item: (-int(item.get("lines") or 0), item["sku"]))


def ensure_warehouse_relation_store() -> tuple[str, str]:
    from .db import ensure_product_schema, ensure_sku_mapping_schema

    url = database_url()
    ensure_sku_mapping_schema(url)
    ensure_product_schema(url)
    return url, load_settings().shop_key


def int_or_none(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def json_list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def localized_text(items: Any, key: str) -> str:
    preferred = ("zh-cn", "en", "fr", "es", "de", "pt-br", "th", "ja", "ko")
    by_language: dict[str, str] = {}
    fallback = ""
    for item in json_list(items):
        if not isinstance(item, dict):
            continue
        text_value = clean_text(item.get(key))
        if not text_value:
            continue
        language = clean_text(item.get("language")).lower()
        if language:
            by_language[language] = text_value
        fallback = fallback or text_value
    for language in preferred:
        if by_language.get(language):
            return by_language[language]
    return fallback


def localized_attributes(items: Any) -> list[dict[str, str]]:
    grouped: dict[tuple[str, str], dict[str, dict[str, str]]] = {}
    fallback_order: list[tuple[str, str]] = []
    for item in json_list(items):
        if not isinstance(item, dict):
            continue
        attr_id = clean_text(item.get("attributeId"))
        value_id = clean_text(item.get("attributeValueId"))
        if not attr_id and not value_id:
            continue
        group_key = (attr_id, value_id)
        if group_key not in grouped:
            grouped[group_key] = {}
            fallback_order.append(group_key)
        language = clean_text(item.get("language")).lower() or "_"
        grouped[group_key][language] = {
            "name": clean_text(item.get("attributeName")),
            "value": clean_text(item.get("attributeValueName")),
        }
    out: list[dict[str, str]] = []
    for group_key in fallback_order:
        candidates = grouped[group_key]
        selected = candidates.get("zh-cn") or candidates.get("en") or next(iter(candidates.values()))
        if selected.get("name") or selected.get("value"):
            out.append(selected)
    return out


def compact_attrs(items: list[dict[str, str]]) -> str:
    parts = []
    for item in items:
        name = clean_text(item.get("name"))
        value = clean_text(item.get("value"))
        if name and value:
            parts.append(f"{name}: {value}")
        elif value:
            parts.append(value)
    return " · ".join(parts)


def summarize_prices(items: Any) -> dict[str, Any]:
    rows = [item for item in json_list(items) if isinstance(item, dict)]
    display: list[str] = []
    min_price: float | None = None
    for item in rows:
        site = clean_text(item.get("site"))
        currency = clean_text(item.get("currency"))
        base = number(item.get("basePrice"))
        special = number(item.get("specialPrice"))
        active = special if special > 0 else base
        if active > 0:
            min_price = active if min_price is None else min(min_price, active)
        label = f"{site} {currency} {active:.2f}".strip()
        if label:
            display.append(label)
    return {"rows": rows, "display": " / ".join(display[:3]), "min": min_price}


def summarize_costs(items: Any) -> dict[str, Any]:
    rows = [item for item in json_list(items) if isinstance(item, dict)]
    display = []
    for item in rows:
        currency = clean_text(item.get("currency"))
        cost = number(item.get("cost"))
        display.append(f"{currency} {cost:.2f}".strip())
    return {"rows": rows, "display": " / ".join(display[:3])}


def summarize_inventory(items: Any) -> dict[str, Any]:
    rows = [item for item in json_list(items) if isinstance(item, dict)]
    total = sum(int_or_none(item.get("inventoryNum")) or 0 for item in rows)
    display = " / ".join(
        f"{clean_text(item.get('warehouseId'))}: {int_or_none(item.get('inventoryNum')) or 0}"
        for item in rows[:4]
    )
    return {"rows": rows, "total": total, "display": display, "warehouses": len(rows)}


def build_product_sku_catalog(conn: psycopg.Connection, shop_key: str) -> dict[str, dict[str, Any]]:
    catalog: dict[str, dict[str, Any]] = {}
    detail_rows = conn.execute(
        """
        SELECT spu_name, spu_shelf_status, category_id, skc_list, last_seen_at
        FROM shein_product_details
        WHERE shop_key = %s
        ORDER BY last_seen_at DESC, spu_name
        """,
        (shop_key,),
    ).fetchall()
    for row in detail_rows:
        spu_name = clean_text(row.get("spu_name"))
        for skc in json_list(row.get("skc_list")):
            if not isinstance(skc, dict):
                continue
            skc_name = clean_text(skc.get("skcName"))
            skc_attrs = localized_attributes(skc.get("skcSalesAttribute"))
            skc_title = localized_text(skc.get("skcTitle"), "title")
            site_status = [item for item in json_list(skc.get("skcSiteShelfStatusList")) if isinstance(item, dict)]
            for sku in json_list(skc.get("skuList")):
                if not isinstance(sku, dict):
                    continue
                sku_code = clean_text(sku.get("skuCode"))
                if not sku_code:
                    continue
                sku_attrs = localized_attributes(sku.get("skuSalesAttributeList"))
                price = summarize_prices(sku.get("priceList"))
                cost = summarize_costs(sku.get("costList"))
                inventory = summarize_inventory(sku.get("inventoryList"))
                catalog[sku_code] = {
                    "skuCode": sku_code,
                    "supplierSku": clean_text(sku.get("supplierSku")),
                    "skcName": skc_name,
                    "spuName": spu_name,
                    "categoryId": clean_text(row.get("category_id")),
                    "spuShelfStatus": int_or_none(row.get("spu_shelf_status")),
                    "skcShelfStatus": int_or_none(skc.get("skcShelfStatus")),
                    "supplierCode": clean_text(skc.get("supplierCode")),
                    "title": skc_title,
                    "imageUrl": clean_text(skc.get("skcMainPicUrl")),
                    "skcAttributes": skc_attrs,
                    "skuAttributes": sku_attrs,
                    "attributeSummary": compact_attrs([*skc_attrs, *sku_attrs]),
                    "priceList": price["rows"],
                    "priceText": price["display"],
                    "minPrice": price["min"],
                    "costList": cost["rows"],
                    "costText": cost["display"],
                    "inventoryList": inventory["rows"],
                    "inventoryText": inventory["display"],
                    "inventoryTotal": inventory["total"],
                    "inventoryWarehouses": inventory["warehouses"],
                    "siteShelfStatusList": site_status,
                    "productLastSeenAt": row.get("last_seen_at"),
                }
    product_rows = conn.execute(
        """
        SELECT spu_name, skc_name, sku_code_list, last_seen_at
        FROM shein_products
        WHERE shop_key = %s
        ORDER BY last_seen_at DESC, spu_name, skc_name
        """,
        (shop_key,),
    ).fetchall()
    for row in product_rows:
        spu_name = clean_text(row.get("spu_name"))
        skc_name = clean_text(row.get("skc_name"))
        for sku_code in [clean_text(item) for item in json_list(row.get("sku_code_list")) if clean_text(item)]:
            catalog.setdefault(
                sku_code,
                {
                    "skuCode": sku_code,
                    "supplierSku": "",
                    "skcName": skc_name,
                    "spuName": spu_name,
                    "categoryId": "",
                    "spuShelfStatus": None,
                    "skcShelfStatus": None,
                    "supplierCode": "",
                    "title": "",
                    "imageUrl": "",
                    "skcAttributes": [],
                    "skuAttributes": [],
                    "attributeSummary": "",
                    "priceList": [],
                    "priceText": "",
                    "minPrice": None,
                    "costList": [],
                    "costText": "",
                    "inventoryList": [],
                    "inventoryText": "",
                    "inventoryTotal": 0,
                    "inventoryWarehouses": 0,
                    "siteShelfStatusList": [],
                    "productLastSeenAt": row.get("last_seen_at"),
                },
            )
    return catalog


def serialize_relation_datetime(value: Any) -> str:
    return value.strftime("%Y-%m-%d %H:%M") if hasattr(value, "strftime") else clean_text(value)


INVENTORY_CALCULATION_MODES = {"direct_amount", "quantity_x_unit"}
INVENTORY_DEFAULT_TEMPLATE_NAME = "标准库存成本模板"
INVENTORY_DEFAULT_TEMPLATE = [
    {
        "name": "出口代理费用",
        "currency": "RMB",
        "sortOrder": 10,
        "types": [
            {"name": "拖车费-起运港", "calculationMode": "quantity_x_unit", "sortOrder": 10},
            {"name": "港杂物费", "calculationMode": "quantity_x_unit", "sortOrder": 20},
        ],
    },
    {
        "name": "海运进口费用",
        "currency": "USD",
        "sortOrder": 20,
        "types": [
            {"name": "税金", "calculationMode": "quantity_x_unit", "sortOrder": 10},
            {"name": "尾程派送费", "calculationMode": "quantity_x_unit", "sortOrder": 20},
            {"name": "美国反恐舱单费", "calculationMode": "quantity_x_unit", "sortOrder": 30},
            {"name": "海运费", "calculationMode": "quantity_x_unit", "sortOrder": 40},
        ],
    },
    {
        "name": "入仓费用",
        "currency": "USD",
        "sortOrder": 30,
        "types": [
            {"name": "待时费", "calculationMode": "quantity_x_unit", "sortOrder": 10},
            {"name": "卸货费", "calculationMode": "quantity_x_unit", "sortOrder": 20},
        ],
    },
]


def ensure_inventory_store() -> tuple[str, str]:
    from .db import ensure_inventory_schema, ensure_product_schema, ensure_sku_mapping_schema

    url = database_url()
    ensure_inventory_schema(url)
    ensure_sku_mapping_schema(url)
    ensure_product_schema(url)
    return url, load_settings().shop_key


def inventory_note(value: Any) -> str | None:
    text = clean_text(value).strip()
    if len(text) > 1000:
        raise HTTPException(status_code=400, detail="note is too long")
    return text or None


def inventory_decimal(value: Any, field_name: str) -> Decimal | None:
    return sku_mapping_decimal(value, field_name)


def inventory_number_or_none(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return float(value)


def inventory_datetime(value: Any) -> str:
    return value.strftime("%Y-%m-%d %H:%M") if hasattr(value, "strftime") else clean_text(value)


def inventory_name(value: Any, field_name: str) -> str:
    text = clean_text(value).strip()
    if not text:
        raise HTTPException(status_code=400, detail=f"{field_name} is required")
    if len(text) > 120:
        raise HTTPException(status_code=400, detail=f"{field_name} is too long")
    return text


def inventory_currency(value: Any) -> str:
    currency = clean_text(value).strip().upper()
    if not re.fullmatch(r"[A-Z][A-Z0-9]{1,7}", currency):
        raise HTTPException(status_code=400, detail="currency must be an uppercase code")
    return currency


def inventory_sort_order(value: Any, default: int = 0) -> int:
    if value in (None, ""):
        return default
    try:
        order = int(value)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="sortOrder must be an integer")
    if abs(order) > 1_000_000:
        raise HTTPException(status_code=400, detail="sortOrder is out of range")
    return order


def inventory_calculation_mode(value: Any) -> str:
    mode = clean_text(value) or "quantity_x_unit"
    if mode not in INVENTORY_CALCULATION_MODES:
        raise HTTPException(status_code=400, detail="invalid calculationMode")
    return mode


def ensure_inventory_template(conn: psycopg.Connection, shop_key: str) -> None:
    initialized = conn.execute(
        """
        INSERT INTO inventory_template_settings (shop_key)
        VALUES (%s)
        ON CONFLICT (shop_key) DO NOTHING
        RETURNING shop_key
        """,
        (shop_key,),
    ).fetchone()
    if not initialized:
        return
    template = conn.execute(
        """
        INSERT INTO inventory_cost_templates (shop_key, name)
        VALUES (%s, %s)
        RETURNING id
        """,
        (shop_key, INVENTORY_DEFAULT_TEMPLATE_NAME),
    ).fetchone()
    template_id = int(template.get("id"))
    for category in INVENTORY_DEFAULT_TEMPLATE:
        row = conn.execute(
            """
            INSERT INTO inventory_cost_categories (
                shop_key, template_id, name, currency, sort_order
            )
            VALUES (%s, %s, %s, %s, %s)
            RETURNING id
            """,
            (
                shop_key,
                template_id,
                category["name"],
                category["currency"],
                category["sortOrder"],
            ),
        ).fetchone()
        category_id = int(row.get("id"))
        for cost_type in category["types"]:
            conn.execute(
                """
                INSERT INTO inventory_cost_types (category_id, name, calculation_mode, sort_order)
                VALUES (%s, %s, %s, %s)
                """,
                (
                    category_id,
                    cost_type["name"],
                    cost_type["calculationMode"],
                    cost_type["sortOrder"],
                ),
            )


def fetch_inventory_templates(conn: psycopg.Connection, shop_key: str) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT tm.id AS template_id, tm.name AS template_name,
               tm.created_at AS template_created_at, tm.updated_at AS template_updated_at,
               c.id AS category_id, c.name AS category_name, c.currency,
               c.sort_order AS category_sort_order,
               t.id AS cost_type_id, t.name AS cost_type_name,
               t.calculation_mode, t.sort_order AS cost_type_sort_order
        FROM inventory_cost_templates tm
        LEFT JOIN inventory_cost_categories c ON c.template_id = tm.id
        LEFT JOIN inventory_cost_types t ON t.category_id = c.id
        WHERE tm.shop_key = %s
        ORDER BY tm.updated_at DESC, tm.id, c.sort_order, c.id, t.sort_order, t.id
        """,
        (shop_key,),
    ).fetchall()
    templates: list[dict[str, Any]] = []
    templates_by_id: dict[int, dict[str, Any]] = {}
    categories_by_id: dict[int, dict[str, Any]] = {}
    for row in rows:
        template_id = int(row.get("template_id"))
        template = templates_by_id.get(template_id)
        if template is None:
            template = {
                "id": template_id,
                "name": clean_text(row.get("template_name")),
                "createdAt": inventory_datetime(row.get("template_created_at")),
                "updatedAt": inventory_datetime(row.get("template_updated_at")),
                "categories": [],
            }
            templates_by_id[template_id] = template
            templates.append(template)
        if row.get("category_id") is None:
            continue
        category_id = int(row.get("category_id"))
        category = categories_by_id.get(category_id)
        if category is None:
            category = {
                "id": category_id,
                "templateId": template_id,
                "name": clean_text(row.get("category_name")),
                "currency": clean_text(row.get("currency")),
                "sortOrder": int(row.get("category_sort_order") or 0),
                "types": [],
            }
            categories_by_id[category_id] = category
            template["categories"].append(category)
        if row.get("cost_type_id") is not None:
            category["types"].append(
                {
                    "id": int(row.get("cost_type_id")),
                    "categoryId": category_id,
                    "name": clean_text(row.get("cost_type_name")),
                    "calculationMode": clean_text(row.get("calculation_mode")),
                    "sortOrder": int(row.get("cost_type_sort_order") or 0),
                }
            )
    return templates


def inventory_template_response(conn: psycopg.Connection, shop_key: str) -> JSONResponse:
    return JSONResponse({"templates": fetch_inventory_templates(conn, shop_key)})



def inventory_warehouse_references(conn: psycopg.Connection, shop_key: str) -> list[dict[str, Any]]:
    warehouse_rows = conn.execute(
        """
        SELECT warehouse_sku, enabled, updated_at
        FROM shein_warehouse_skus
        WHERE shop_key = %s
        ORDER BY enabled DESC, warehouse_sku
        """,
        (shop_key,),
    ).fetchall()
    mapping_rows = conn.execute(
        """
        SELECT warehouse_sku, shein_sku
        FROM shein_sku_mappings
        WHERE shop_key = %s AND enabled = true
        ORDER BY warehouse_sku, shein_sku
        """,
        (shop_key,),
    ).fetchall()
    catalog = build_product_sku_catalog(conn, shop_key)
    refs: dict[str, dict[str, Any]] = {}
    for row in warehouse_rows:
        warehouse_sku = clean_text(row.get("warehouse_sku"))
        if not warehouse_sku:
            continue
        refs[warehouse_sku] = {
            "warehouseSku": warehouse_sku,
            "enabled": bool(row.get("enabled")),
            "updatedAt": inventory_datetime(row.get("updated_at")),
            "items": [],
        }
    for row in mapping_rows:
        warehouse_sku = clean_text(row.get("warehouse_sku"))
        shein_sku = clean_text(row.get("shein_sku"))
        if not warehouse_sku or not shein_sku:
            continue
        product = catalog.get(shein_sku, {})
        refs.setdefault(
            warehouse_sku,
            {"warehouseSku": warehouse_sku, "enabled": True, "updatedAt": "", "items": []},
        )
        refs[warehouse_sku]["items"].append(
            {
                "sheinSku": shein_sku,
                "imageUrl": clean_text(product.get("imageUrl")),
                "title": clean_text(product.get("title") or product.get("attributeSummary") or shein_sku),
                "skcName": clean_text(product.get("skcName")),
                "spuName": clean_text(product.get("spuName")),
                "supplierSku": clean_text(product.get("supplierSku")),
                "supplierCode": clean_text(product.get("supplierCode")),
            }
        )
    return sorted(refs.values(), key=lambda item: item["warehouseSku"])


def inventory_cost_totals(rows: list[dict[str, Any]]) -> dict[str, float]:
    totals: dict[str, float] = {}
    for row in rows:
        currency = clean_text(row.get("currency"))
        if not currency:
            continue
        totals[currency] = round(totals.get(currency, 0.0) + float(row.get("amount") or 0), 2)
    return totals


def inventory_ticket_cost_totals(
    conn: psycopg.Connection,
    *,
    shop_key: str,
    ticket_ids: list[int],
) -> dict[int, dict[str, float]]:
    if not ticket_ids:
        return {}
    rows = conn.execute(
        """
        SELECT e.inventory_ticket_id, c.currency, SUM(COALESCE(e.amount, 0)) AS amount
        FROM inventory_cost_entries e
        JOIN inventory_cost_types t ON t.id = e.cost_type_id
        JOIN inventory_cost_categories c ON c.id = t.category_id
        JOIN inventory_tickets i ON i.id = e.inventory_ticket_id
        WHERE i.shop_key = %s AND e.inventory_ticket_id = ANY(%s)
        GROUP BY e.inventory_ticket_id, c.currency
        """,
        (shop_key, ticket_ids),
    ).fetchall()
    totals: dict[int, dict[str, float]] = {}
    for row in rows:
        ticket_id = int(row.get("inventory_ticket_id"))
        totals.setdefault(ticket_id, {})[clean_text(row.get("currency"))] = round(
            float(row.get("amount") or 0), 2
        )
    return totals


def serialize_inventory_ticket_summary(
    row: dict[str, Any],
    *,
    cost_totals: dict[int, dict[str, float]],
) -> dict[str, Any]:
    ticket_id = int(row.get("id"))
    return {
        "ticketId": ticket_id,
        "ticketNo": clean_text(row.get("ticket_no")),
        "costTemplateId": row.get("cost_template_id"),
        "costTemplateName": clean_text(row.get("cost_template_name")),
        "note": clean_text(row.get("note")),
        "lineCount": int(row.get("line_count") or 0),
        "warehouseSkuCount": int(row.get("warehouse_sku_count") or 0),
        "quantity": inventory_number_or_none(row.get("quantity")) or 0,
        "inventoryAmount": inventory_number_or_none(row.get("inventory_amount")) or 0,
        "costTotals": cost_totals.get(ticket_id, {}),
        "createdAt": inventory_datetime(row.get("created_at")),
        "updatedAt": inventory_datetime(row.get("updated_at")),
    }


def fetch_inventory_ticket_detail(
    conn: psycopg.Connection,
    *,
    shop_key: str,
    ticket_id: int,
) -> dict[str, Any] | None:
    ticket = conn.execute(
        """
        SELECT i.id, i.ticket_no, i.cost_template_id, tm.name AS cost_template_name,
               i.note, i.created_at, i.updated_at
        FROM inventory_tickets i
        LEFT JOIN inventory_cost_templates tm ON tm.id = i.cost_template_id
        WHERE i.id = %s AND i.shop_key = %s
        """,
        (ticket_id, shop_key),
    ).fetchone()
    if not ticket:
        return None
    line_rows = conn.execute(
        """
        SELECT id, warehouse_sku, quantity, unit_price, amount, note, created_at, updated_at
        FROM inventory_ticket_lines
        WHERE inventory_ticket_id = %s
        ORDER BY id
        """,
        (ticket_id,),
    ).fetchall()
    cost_rows = conn.execute(
        """
        SELECT e.id, e.cost_type_id, e.quantity, e.unit_price, e.amount, e.note,
               t.category_id, t.name AS cost_type_name, t.calculation_mode,
               c.name AS category_name, c.currency
        FROM inventory_cost_entries e
        JOIN inventory_cost_types t ON t.id = e.cost_type_id
        JOIN inventory_cost_categories c ON c.id = t.category_id
        WHERE e.inventory_ticket_id = %s AND c.shop_key = %s
        ORDER BY c.sort_order, c.id, t.sort_order, t.id
        """,
        (ticket_id, shop_key),
    ).fetchall()
    lines = [
        {
            "id": row.get("id"),
            "warehouseSku": clean_text(row.get("warehouse_sku")),
            "quantity": inventory_number_or_none(row.get("quantity")) or 0,
            "unitPrice": inventory_number_or_none(row.get("unit_price")),
            "amount": inventory_number_or_none(row.get("amount")),
            "note": clean_text(row.get("note")),
        }
        for row in line_rows
    ]
    costs = [
        {
            "id": row.get("id"),
            "costTypeId": row.get("cost_type_id"),
            "categoryId": row.get("category_id"),
            "categoryName": clean_text(row.get("category_name")),
            "costTypeName": clean_text(row.get("cost_type_name")),
            "calculationMode": clean_text(row.get("calculation_mode")),
            "currency": clean_text(row.get("currency")),
            "quantity": inventory_number_or_none(row.get("quantity")),
            "unitPrice": inventory_number_or_none(row.get("unit_price")),
            "amount": inventory_number_or_none(row.get("amount")),
            "note": clean_text(row.get("note")),
        }
        for row in cost_rows
    ]
    return {
        "ticketId": int(ticket.get("id")),
        "ticketNo": clean_text(ticket.get("ticket_no")),
        "costTemplateId": ticket.get("cost_template_id"),
        "costTemplateName": clean_text(ticket.get("cost_template_name")),
        "note": clean_text(ticket.get("note")),
        "lines": lines,
        "costs": costs,
        "costTotals": inventory_cost_totals(cost_rows),
        "createdAt": inventory_datetime(ticket.get("created_at")),
        "updatedAt": inventory_datetime(ticket.get("updated_at")),
    }


def normalize_inventory_lines(lines_payload: Any) -> list[dict[str, Any]]:
    if not isinstance(lines_payload, list) or not lines_payload:
        raise HTTPException(status_code=400, detail="at least one inventory line is required")
    lines: list[dict[str, Any]] = []
    seen: set[str] = set()
    for index, payload in enumerate(lines_payload):
        if not isinstance(payload, dict):
            raise HTTPException(status_code=400, detail=f"lines[{index}] must be an object")
        warehouse_sku = sku_mapping_text(
            payload_value(payload, "warehouseSku", "warehouse_sku"),
            f"lines[{index}].warehouseSku",
        )
        if warehouse_sku in seen:
            raise HTTPException(status_code=400, detail=f"duplicate warehouse SKU: {warehouse_sku}")
        seen.add(warehouse_sku)
        quantity = inventory_decimal(payload_value(payload, "quantity", "qty"), f"lines[{index}].quantity")
        quantity = quantity if quantity is not None else Decimal("0")
        unit_price = inventory_decimal(
            payload_value(payload, "unitPrice", "unit_price"),
            f"lines[{index}].unitPrice",
        )
        amount = inventory_decimal(payload_value(payload, "amount"), f"lines[{index}].amount")
        if amount is None and unit_price is not None:
            amount = quantity * unit_price
        lines.append(
            {
                "warehouse_sku": warehouse_sku,
                "quantity": quantity,
                "unit_price": unit_price,
                "amount": amount,
                "note": inventory_note(payload_value(payload, "note")),
            }
        )
    return lines


def inventory_ticket_template_id(
    conn: psycopg.Connection,
    *,
    shop_key: str,
    value: Any,
) -> int | None:
    if value in (None, ""):
        return None
    try:
        template_id = int(value)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="costTemplateId must be an integer")
    row = conn.execute(
        """
        SELECT id FROM inventory_cost_templates
        WHERE id = %s AND shop_key = %s
        """,
        (template_id, shop_key),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=400, detail="cost template not found")
    return template_id


def normalize_inventory_costs(
    conn: psycopg.Connection,
    *,
    shop_key: str,
    cost_template_id: int | None,
    costs_payload: Any,
) -> list[dict[str, Any]]:
    if costs_payload in (None, ""):
        return []
    if not isinstance(costs_payload, list):
        raise HTTPException(status_code=400, detail="costs must be an array")
    if cost_template_id is None:
        if costs_payload:
            raise HTTPException(status_code=400, detail="blank template cannot contain costs")
        return []
    type_ids: list[int] = []
    raw_by_id: dict[int, dict[str, Any]] = {}
    for index, payload in enumerate(costs_payload):
        if not isinstance(payload, dict):
            raise HTTPException(status_code=400, detail=f"costs[{index}] must be an object")
        try:
            cost_type_id = int(payload_value(payload, "costTypeId", "cost_type_id"))
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail=f"costs[{index}].costTypeId is required")
        if cost_type_id in raw_by_id:
            raise HTTPException(status_code=400, detail=f"duplicate costTypeId: {cost_type_id}")
        type_ids.append(cost_type_id)
        raw_by_id[cost_type_id] = payload
    if not type_ids:
        return []
    type_rows = conn.execute(
        """
        SELECT t.id, t.calculation_mode
        FROM inventory_cost_types t
        JOIN inventory_cost_categories c ON c.id = t.category_id
        JOIN inventory_cost_templates tm ON tm.id = c.template_id
        WHERE tm.shop_key = %s AND tm.id = %s AND t.id = ANY(%s)
        """,
        (shop_key, cost_template_id, type_ids),
    ).fetchall()
    modes = {int(row.get("id")): clean_text(row.get("calculation_mode")) for row in type_rows}
    missing = sorted(set(type_ids) - set(modes))
    if missing:
        raise HTTPException(status_code=400, detail=f"unknown costTypeId: {missing[0]}")
    costs: list[dict[str, Any]] = []
    for index, cost_type_id in enumerate(type_ids):
        payload = raw_by_id[cost_type_id]
        note = inventory_note(payload_value(payload, "note"))
        quantity = inventory_decimal(
            payload_value(payload, "quantity", "qty"),
            f"costs[{index}].quantity",
        )
        unit_price = inventory_decimal(
            payload_value(payload, "unitPrice", "unit_price"),
            f"costs[{index}].unitPrice",
        )
        amount = inventory_decimal(payload_value(payload, "amount"), f"costs[{index}].amount")
        mode = modes[cost_type_id]
        if mode == "direct_amount":
            quantity = None
            unit_price = None
        elif quantity is None and unit_price is None:
            amount = None
        elif quantity is None or unit_price is None:
            raise HTTPException(
                status_code=400,
                detail=f"costs[{index}] requires both quantity and unitPrice",
            )
        else:
            amount = quantity * unit_price
        if quantity is None and unit_price is None and amount is None and note is None:
            continue
        costs.append(
            {
                "cost_type_id": cost_type_id,
                "quantity": quantity,
                "unit_price": unit_price,
                "amount": amount,
                "note": note,
            }
        )
    return costs


def save_inventory_ticket(
    conn: psycopg.Connection,
    *,
    shop_key: str,
    payload: dict[str, Any],
    ticket_id: int | None = None,
) -> int:
    ticket_no = inventory_name(
        payload_value(payload, "ticketNo", "ticket_no"),
        "ticketNo",
    )
    note = inventory_note(payload_value(payload, "note"))
    cost_template_id = inventory_ticket_template_id(
        conn,
        shop_key=shop_key,
        value=payload_value(payload, "costTemplateId", "cost_template_id"),
    )
    lines = normalize_inventory_lines(payload_value(payload, "lines"))
    costs = normalize_inventory_costs(
        conn,
        shop_key=shop_key,
        cost_template_id=cost_template_id,
        costs_payload=payload_value(payload, "costs"),
    )
    if ticket_id is None:
        ticket = conn.execute(
            """
            INSERT INTO inventory_tickets (shop_key, ticket_no, cost_template_id, note)
            VALUES (%s, %s, %s, %s)
            RETURNING id
            """,
            (shop_key, ticket_no, cost_template_id, note),
        ).fetchone()
        ticket_id = int(ticket.get("id"))
    else:
        ticket = conn.execute(
            """
            UPDATE inventory_tickets
            SET ticket_no = %s, cost_template_id = %s, note = %s, updated_at = now()
            WHERE id = %s AND shop_key = %s
            RETURNING id
            """,
            (ticket_no, cost_template_id, note, ticket_id, shop_key),
        ).fetchone()
        if not ticket:
            raise HTTPException(status_code=404, detail="inventory ticket not found")
        conn.execute(
            "DELETE FROM inventory_ticket_lines WHERE inventory_ticket_id = %s",
            (ticket_id,),
        )
        conn.execute(
            "DELETE FROM inventory_cost_entries WHERE inventory_ticket_id = %s",
            (ticket_id,),
        )
    for line in lines:
        conn.execute(
            """
            INSERT INTO inventory_ticket_lines (
                inventory_ticket_id, warehouse_sku, quantity, unit_price, amount, note
            )
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (
                ticket_id,
                line["warehouse_sku"],
                line["quantity"],
                line["unit_price"],
                line["amount"],
                line["note"],
            ),
        )
    for cost in costs:
        conn.execute(
            """
            INSERT INTO inventory_cost_entries (
                inventory_ticket_id, cost_type_id, quantity, unit_price, amount, note
            )
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (
                ticket_id,
                cost["cost_type_id"],
                cost["quantity"],
                cost["unit_price"],
                cost["amount"],
                cost["note"],
            ),
        )
    return ticket_id


@app.get("/api/inventory")
def api_inventory(
    q: str | None = None,
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_INVENTORY)
    url, shop_key = ensure_inventory_store()
    filters = ["t.shop_key = %s"]
    params: list[Any] = [shop_key]
    if q:
        term = f"%{q.strip()}%"
        filters.append(
            """(
                t.ticket_no ILIKE %s
                OR COALESCE(t.note, '') ILIKE %s
                OR EXISTS (
                    SELECT 1 FROM inventory_ticket_lines ql
                    WHERE ql.inventory_ticket_id = t.id AND ql.warehouse_sku ILIKE %s
                )
            )"""
        )
        params.extend([term, term, term])
    where_sql = " AND ".join(filters)
    params.append(600)
    with psycopg.connect(url, row_factory=dict_row) as conn:
        ensure_inventory_template(conn, shop_key)
        ticket_rows = conn.execute(
            f"""
            SELECT t.id, t.ticket_no, t.cost_template_id,
                   tm.name AS cost_template_name,
                   t.note, t.created_at, t.updated_at,
                   COUNT(l.id) AS line_count,
                   COUNT(DISTINCT l.warehouse_sku) AS warehouse_sku_count,
                   COALESCE(SUM(l.quantity), 0) AS quantity,
                   COALESCE(SUM(l.amount), 0) AS inventory_amount
            FROM inventory_tickets t
            LEFT JOIN inventory_cost_templates tm ON tm.id = t.cost_template_id
            LEFT JOIN inventory_ticket_lines l ON l.inventory_ticket_id = t.id
            WHERE {where_sql}
            GROUP BY t.id, tm.name
            ORDER BY t.updated_at DESC, t.id DESC
            LIMIT %s
            """,
            params,
        ).fetchall()
        ticket_ids = [int(row.get("id")) for row in ticket_rows]
        cost_totals = inventory_ticket_cost_totals(conn, shop_key=shop_key, ticket_ids=ticket_ids)
        templates = fetch_inventory_templates(conn, shop_key)
        warehouse_refs = inventory_warehouse_references(conn, shop_key)
    tickets = [
        serialize_inventory_ticket_summary(row, cost_totals=cost_totals)
        for row in ticket_rows
    ]
    total_costs: dict[str, float] = {}
    for totals in cost_totals.values():
        for currency, amount in totals.items():
            total_costs[currency] = round(total_costs.get(currency, 0.0) + amount, 2)
    return JSONResponse(
        {
            "shop": shop_key,
            "summary": {
                "tickets": len(tickets),
                "lines": sum(ticket["lineCount"] for ticket in tickets),
                "warehouseSkus": len(
                    {
                        item["warehouseSku"]
                        for item in warehouse_refs
                        if item.get("warehouseSku")
                    }
                ),
                "quantity": round(sum(float(ticket["quantity"]) for ticket in tickets), 3),
                "inventoryAmount": round(
                    sum(float(ticket["inventoryAmount"]) for ticket in tickets),
                    2,
                ),
                "costTotals": total_costs,
                "limit": 600,
            },
            "tickets": tickets,
            "templates": templates,
            "warehouseRefs": warehouse_refs,
        }
    )


@app.get("/api/inventory/{ticket_id}")
def api_inventory_ticket(
    ticket_id: int,
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_INVENTORY)
    url, shop_key = ensure_inventory_store()
    with psycopg.connect(url, row_factory=dict_row) as conn:
        ensure_inventory_template(conn, shop_key)
        ticket = fetch_inventory_ticket_detail(conn, shop_key=shop_key, ticket_id=ticket_id)
        templates = fetch_inventory_templates(conn, shop_key)
    if ticket is None:
        raise HTTPException(status_code=404, detail="inventory ticket not found")
    return JSONResponse({"ticket": ticket, "templates": templates})


@app.post("/api/inventory")
def api_create_inventory_ticket(
    payload: dict[str, Any] = Body(...),
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_INVENTORY)
    url, shop_key = ensure_inventory_store()
    try:
        with psycopg.connect(url, row_factory=dict_row) as conn:
            ensure_inventory_template(conn, shop_key)
            ticket_id = save_inventory_ticket(conn, shop_key=shop_key, payload=payload)
            ticket = fetch_inventory_ticket_detail(conn, shop_key=shop_key, ticket_id=ticket_id)
    except psycopg.errors.UniqueViolation:
        raise HTTPException(status_code=409, detail="inventory ticket number or warehouse SKU already exists")
    return JSONResponse({"saved": True, "ticket": ticket}, status_code=201)


@app.put("/api/inventory/{ticket_id}")
def api_update_inventory_ticket(
    ticket_id: int,
    payload: dict[str, Any] = Body(...),
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_INVENTORY)
    url, shop_key = ensure_inventory_store()
    try:
        with psycopg.connect(url, row_factory=dict_row) as conn:
            ensure_inventory_template(conn, shop_key)
            save_inventory_ticket(
                conn,
                shop_key=shop_key,
                payload=payload,
                ticket_id=ticket_id,
            )
            ticket = fetch_inventory_ticket_detail(conn, shop_key=shop_key, ticket_id=ticket_id)
    except psycopg.errors.UniqueViolation:
        raise HTTPException(status_code=409, detail="inventory ticket number or warehouse SKU already exists")
    return JSONResponse({"saved": True, "ticket": ticket})


@app.delete("/api/inventory/{ticket_id}")
def api_delete_inventory_ticket(
    ticket_id: int,
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_INVENTORY)
    url, shop_key = ensure_inventory_store()
    with psycopg.connect(url, row_factory=dict_row) as conn:
        row = conn.execute(
            """
            DELETE FROM inventory_tickets
            WHERE id = %s AND shop_key = %s
            RETURNING id
            """,
            (ticket_id, shop_key),
        ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="inventory ticket not found")
    return JSONResponse({"deleted": True})


@app.get("/api/inventory-cost-templates")
def api_inventory_cost_templates(
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_COST_TEMPLATES)
    url, shop_key = ensure_inventory_store()
    with psycopg.connect(url, row_factory=dict_row) as conn:
        ensure_inventory_template(conn, shop_key)
        return inventory_template_response(conn, shop_key)


@app.post("/api/inventory-cost-templates")
def api_create_inventory_cost_template(
    payload: dict[str, Any] = Body(...),
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_COST_TEMPLATES)
    url, shop_key = ensure_inventory_store()
    name = inventory_name(payload_value(payload, "name"), "name")
    try:
        with psycopg.connect(url, row_factory=dict_row) as conn:
            ensure_inventory_template(conn, shop_key)
            conn.execute(
                """
                INSERT INTO inventory_cost_templates (shop_key, name)
                VALUES (%s, %s)
                """,
                (shop_key, name),
            )
            return inventory_template_response(conn, shop_key)
    except psycopg.errors.UniqueViolation:
        raise HTTPException(status_code=409, detail="cost template name already exists")


@app.patch("/api/inventory-cost-templates/{template_id}")
def api_update_inventory_cost_template(
    template_id: int,
    payload: dict[str, Any] = Body(...),
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_COST_TEMPLATES)
    url, shop_key = ensure_inventory_store()
    name = inventory_name(payload_value(payload, "name"), "name")
    try:
        with psycopg.connect(url, row_factory=dict_row) as conn:
            updated = conn.execute(
                """
                UPDATE inventory_cost_templates
                SET name = %s, updated_at = now()
                WHERE id = %s AND shop_key = %s
                RETURNING id
                """,
                (name, template_id, shop_key),
            ).fetchone()
            if not updated:
                raise HTTPException(status_code=404, detail="cost template not found")
            return inventory_template_response(conn, shop_key)
    except psycopg.errors.UniqueViolation:
        raise HTTPException(status_code=409, detail="cost template name already exists")


@app.delete("/api/inventory-cost-templates/{template_id}")
def api_delete_inventory_cost_template(
    template_id: int,
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_COST_TEMPLATES)
    url, shop_key = ensure_inventory_store()
    with psycopg.connect(url, row_factory=dict_row) as conn:
        impact = conn.execute(
            """
            SELECT COUNT(DISTINCT i.id) AS tickets,
                   COUNT(DISTINCT e.id) AS entries
            FROM inventory_cost_templates tm
            LEFT JOIN inventory_tickets i ON i.cost_template_id = tm.id
            LEFT JOIN inventory_cost_categories c ON c.template_id = tm.id
            LEFT JOIN inventory_cost_types t ON t.category_id = c.id
            LEFT JOIN inventory_cost_entries e ON e.cost_type_id = t.id
            WHERE tm.id = %s AND tm.shop_key = %s
            """,
            (template_id, shop_key),
        ).fetchone()
        deleted = conn.execute(
            """
            DELETE FROM inventory_cost_templates
            WHERE id = %s AND shop_key = %s
            RETURNING id
            """,
            (template_id, shop_key),
        ).fetchone()
        if not deleted:
            raise HTTPException(status_code=404, detail="cost template not found")
        response = inventory_template_response(conn, shop_key)
        body = json.loads(response.body)
        body["deleted"] = True
        body["affectedTickets"] = int(impact.get("tickets") or 0)
        body["deletedEntries"] = int(impact.get("entries") or 0)
        return JSONResponse(body)


@app.get("/api/inventory-cost-categories")
def api_inventory_cost_categories(
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    return api_inventory_cost_templates(token, shein_pnl_token)


@app.post("/api/inventory-cost-categories")
def api_create_inventory_cost_category(
    payload: dict[str, Any] = Body(...),
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_COST_TEMPLATES)
    url, shop_key = ensure_inventory_store()
    try:
        template_id = int(payload_value(payload, "templateId", "template_id"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="templateId is required")
    name = inventory_name(payload_value(payload, "name"), "name")
    currency = inventory_currency(payload_value(payload, "currency"))
    try:
        with psycopg.connect(url, row_factory=dict_row) as conn:
            template = conn.execute(
                """
                SELECT id FROM inventory_cost_templates
                WHERE id = %s AND shop_key = %s
                """,
                (template_id, shop_key),
            ).fetchone()
            if not template:
                raise HTTPException(status_code=404, detail="cost template not found")
            default_order = conn.execute(
                """
                SELECT COALESCE(MAX(sort_order), 0) + 10 AS sort_order
                FROM inventory_cost_categories
                WHERE template_id = %s
                """,
                (template_id,),
            ).fetchone().get("sort_order")
            conn.execute(
                """
                INSERT INTO inventory_cost_categories (
                    shop_key, template_id, name, currency, sort_order
                )
                VALUES (%s, %s, %s, %s, %s)
                """,
                (
                    shop_key,
                    template_id,
                    name,
                    currency,
                    inventory_sort_order(payload_value(payload, "sortOrder"), int(default_order or 10)),
                ),
            )
            conn.execute(
                "UPDATE inventory_cost_templates SET updated_at = now() WHERE id = %s",
                (template_id,),
            )
            return inventory_template_response(conn, shop_key)
    except psycopg.errors.UniqueViolation:
        raise HTTPException(status_code=409, detail="cost category name already exists in template")


@app.patch("/api/inventory-cost-categories/{category_id}")
def api_update_inventory_cost_category(
    category_id: int,
    payload: dict[str, Any] = Body(...),
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_COST_TEMPLATES)
    url, shop_key = ensure_inventory_store()
    try:
        with psycopg.connect(url, row_factory=dict_row) as conn:
            current = conn.execute(
                """
                SELECT c.id, c.template_id, c.name, c.currency, c.sort_order
                FROM inventory_cost_categories c
                JOIN inventory_cost_templates tm ON tm.id = c.template_id
                WHERE c.id = %s AND tm.shop_key = %s
                """,
                (category_id, shop_key),
            ).fetchone()
            if not current:
                raise HTTPException(status_code=404, detail="cost category not found")
            name = (
                inventory_name(payload.get("name"), "name")
                if "name" in payload
                else clean_text(current.get("name"))
            )
            currency = (
                inventory_currency(payload.get("currency"))
                if "currency" in payload
                else clean_text(current.get("currency"))
            )
            sort_order = inventory_sort_order(
                payload.get("sortOrder"),
                int(current.get("sort_order") or 0),
            )
            conn.execute(
                """
                UPDATE inventory_cost_categories
                SET name = %s, currency = %s, sort_order = %s, updated_at = now()
                WHERE id = %s
                """,
                (name, currency, sort_order, category_id),
            )
            conn.execute(
                "UPDATE inventory_cost_templates SET updated_at = now() WHERE id = %s",
                (current.get("template_id"),),
            )
            return inventory_template_response(conn, shop_key)
    except psycopg.errors.UniqueViolation:
        raise HTTPException(status_code=409, detail="cost category name already exists in template")


@app.delete("/api/inventory-cost-categories/{category_id}")
def api_delete_inventory_cost_category(
    category_id: int,
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_COST_TEMPLATES)
    url, shop_key = ensure_inventory_store()
    with psycopg.connect(url, row_factory=dict_row) as conn:
        impact = conn.execute(
            """
            SELECT c.template_id,
                   COUNT(DISTINCT e.inventory_ticket_id) AS tickets,
                   COUNT(e.id) AS entries
            FROM inventory_cost_categories c
            JOIN inventory_cost_templates tm ON tm.id = c.template_id
            LEFT JOIN inventory_cost_types t ON t.category_id = c.id
            LEFT JOIN inventory_cost_entries e ON e.cost_type_id = t.id
            WHERE c.id = %s AND tm.shop_key = %s
            GROUP BY c.template_id
            """,
            (category_id, shop_key),
        ).fetchone()
        if not impact:
            raise HTTPException(status_code=404, detail="cost category not found")
        conn.execute("DELETE FROM inventory_cost_categories WHERE id = %s", (category_id,))
        conn.execute(
            "UPDATE inventory_cost_templates SET updated_at = now() WHERE id = %s",
            (impact.get("template_id"),),
        )
        response = inventory_template_response(conn, shop_key)
        body = json.loads(response.body)
        body["deleted"] = True
        body["affectedTickets"] = int(impact.get("tickets") or 0)
        body["deletedEntries"] = int(impact.get("entries") or 0)
        return JSONResponse(body)


@app.post("/api/inventory-cost-types")
def api_create_inventory_cost_type(
    payload: dict[str, Any] = Body(...),
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_COST_TEMPLATES)
    url, shop_key = ensure_inventory_store()
    try:
        category_id = int(payload_value(payload, "categoryId", "category_id"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="categoryId is required")
    name = inventory_name(payload_value(payload, "name"), "name")
    mode = inventory_calculation_mode(payload_value(payload, "calculationMode", "calculation_mode"))
    try:
        with psycopg.connect(url, row_factory=dict_row) as conn:
            category = conn.execute(
                """
                SELECT c.id, c.template_id
                FROM inventory_cost_categories c
                JOIN inventory_cost_templates tm ON tm.id = c.template_id
                WHERE c.id = %s AND tm.shop_key = %s
                """,
                (category_id, shop_key),
            ).fetchone()
            if not category:
                raise HTTPException(status_code=404, detail="cost category not found")
            default_order = conn.execute(
                """
                SELECT COALESCE(MAX(sort_order), 0) + 10 AS sort_order
                FROM inventory_cost_types
                WHERE category_id = %s
                """,
                (category_id,),
            ).fetchone().get("sort_order")
            conn.execute(
                """
                INSERT INTO inventory_cost_types (category_id, name, calculation_mode, sort_order)
                VALUES (%s, %s, %s, %s)
                """,
                (
                    category_id,
                    name,
                    mode,
                    inventory_sort_order(payload_value(payload, "sortOrder"), int(default_order or 10)),
                ),
            )
            conn.execute(
                "UPDATE inventory_cost_templates SET updated_at = now() WHERE id = %s",
                (category.get("template_id"),),
            )
            return inventory_template_response(conn, shop_key)
    except psycopg.errors.UniqueViolation:
        raise HTTPException(status_code=409, detail="cost type name already exists in category")


@app.patch("/api/inventory-cost-types/{cost_type_id}")
def api_update_inventory_cost_type(
    cost_type_id: int,
    payload: dict[str, Any] = Body(...),
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_COST_TEMPLATES)
    url, shop_key = ensure_inventory_store()
    try:
        with psycopg.connect(url, row_factory=dict_row) as conn:
            current = conn.execute(
                """
                SELECT t.id, c.template_id, t.name, t.calculation_mode, t.sort_order
                FROM inventory_cost_types t
                JOIN inventory_cost_categories c ON c.id = t.category_id
                JOIN inventory_cost_templates tm ON tm.id = c.template_id
                WHERE t.id = %s AND tm.shop_key = %s
                """,
                (cost_type_id, shop_key),
            ).fetchone()
            if not current:
                raise HTTPException(status_code=404, detail="cost type not found")
            name = (
                inventory_name(payload.get("name"), "name")
                if "name" in payload
                else clean_text(current.get("name"))
            )
            mode = (
                inventory_calculation_mode(payload.get("calculationMode"))
                if "calculationMode" in payload
                else clean_text(current.get("calculation_mode"))
            )
            sort_order = inventory_sort_order(
                payload.get("sortOrder"),
                int(current.get("sort_order") or 0),
            )
            previous_mode = clean_text(current.get("calculation_mode"))
            if previous_mode != mode and mode == "direct_amount":
                conn.execute(
                    """
                    UPDATE inventory_cost_entries
                    SET quantity = NULL, unit_price = NULL, updated_at = now()
                    WHERE cost_type_id = %s
                    """,
                    (cost_type_id,),
                )
            elif previous_mode != mode:
                conn.execute(
                    """
                    UPDATE inventory_cost_entries
                    SET quantity = CASE WHEN amount IS NULL THEN NULL ELSE 1 END,
                        unit_price = amount,
                        updated_at = now()
                    WHERE cost_type_id = %s
                    """,
                    (cost_type_id,),
                )
            conn.execute(
                """
                UPDATE inventory_cost_types
                SET name = %s, calculation_mode = %s, sort_order = %s, updated_at = now()
                WHERE id = %s
                """,
                (name, mode, sort_order, cost_type_id),
            )
            conn.execute(
                "UPDATE inventory_cost_templates SET updated_at = now() WHERE id = %s",
                (current.get("template_id"),),
            )
            return inventory_template_response(conn, shop_key)
    except psycopg.errors.UniqueViolation:
        raise HTTPException(status_code=409, detail="cost type name already exists in category")


@app.delete("/api/inventory-cost-types/{cost_type_id}")
def api_delete_inventory_cost_type(
    cost_type_id: int,
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_COST_TEMPLATES)
    url, shop_key = ensure_inventory_store()
    with psycopg.connect(url, row_factory=dict_row) as conn:
        impact = conn.execute(
            """
            SELECT c.template_id,
                   COUNT(DISTINCT e.inventory_ticket_id) AS tickets,
                   COUNT(e.id) AS entries
            FROM inventory_cost_types t
            JOIN inventory_cost_categories c ON c.id = t.category_id
            JOIN inventory_cost_templates tm ON tm.id = c.template_id
            LEFT JOIN inventory_cost_entries e ON e.cost_type_id = t.id
            WHERE t.id = %s AND tm.shop_key = %s
            GROUP BY c.template_id
            """,
            (cost_type_id, shop_key),
        ).fetchone()
        if not impact:
            raise HTTPException(status_code=404, detail="cost type not found")
        conn.execute("DELETE FROM inventory_cost_types WHERE id = %s", (cost_type_id,))
        conn.execute(
            "UPDATE inventory_cost_templates SET updated_at = now() WHERE id = %s",
            (impact.get("template_id"),),
        )
        response = inventory_template_response(conn, shop_key)
        body = json.loads(response.body)
        body["deleted"] = True
        body["affectedTickets"] = int(impact.get("tickets") or 0)
        body["deletedEntries"] = int(impact.get("entries") or 0)
        return JSONResponse(body)



def build_relation_row(
    *,
    shop_key: str,
    sku_code: str,
    product: dict[str, Any] | None,
    mapping: dict[str, Any] | None,
    default_warehouse_sku: str,
) -> dict[str, Any]:
    status = "mapped" if mapping else "default" if default_warehouse_sku else "unmapped"
    if mapping and product is None:
        status = "missing_product"
    warehouse_sku = clean_text(mapping.get("warehouse_sku")) if mapping else ""
    product = product or {}
    return {
        "shopKey": shop_key,
        "mappingId": mapping.get("id") if mapping else None,
        "mappingStatus": status,
        "warehouseSku": warehouse_sku,
        "effectiveWarehouseSku": warehouse_sku or default_warehouse_sku,
        "warehouseQty": float(mapping.get("warehouse_qty") or 0) if mapping else None,
        "skuGroup": clean_text(mapping.get("sku_group")) if mapping else "",
        "note": clean_text(mapping.get("note")) if mapping else "",
        "enabled": bool(mapping.get("enabled")) if mapping else False,
        "mappingUpdatedAt": serialize_relation_datetime(mapping.get("updated_at")) if mapping else "",
        "skuCode": sku_code,
        "supplierSku": clean_text(product.get("supplierSku")),
        "skcName": clean_text(product.get("skcName")),
        "spuName": clean_text(product.get("spuName")),
        "categoryId": clean_text(product.get("categoryId")),
        "spuShelfStatus": product.get("spuShelfStatus"),
        "skcShelfStatus": product.get("skcShelfStatus"),
        "supplierCode": clean_text(product.get("supplierCode")),
        "title": clean_text(product.get("title")),
        "imageUrl": clean_text(product.get("imageUrl")),
        "attributeSummary": clean_text(product.get("attributeSummary")),
        "skcAttributes": product.get("skcAttributes") or [],
        "skuAttributes": product.get("skuAttributes") or [],
        "priceText": clean_text(product.get("priceText")),
        "minPrice": product.get("minPrice"),
        "priceList": product.get("priceList") or [],
        "costText": clean_text(product.get("costText")),
        "costList": product.get("costList") or [],
        "inventoryText": clean_text(product.get("inventoryText")),
        "inventoryTotal": int(product.get("inventoryTotal") or 0),
        "inventoryWarehouses": int(product.get("inventoryWarehouses") or 0),
        "inventoryList": product.get("inventoryList") or [],
        "siteShelfStatusList": product.get("siteShelfStatusList") or [],
        "productLastSeenAt": serialize_relation_datetime(product.get("productLastSeenAt")),
    }


def relation_group_summary(rows: list[dict[str, Any]], key: str) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        value = clean_text(row.get(key))
        if not value:
            continue
        groups.setdefault(value, []).append(row)
    out = []
    for value, items in groups.items():
        out.append(
            {
                "key": value,
                "skuCount": len({item["skuCode"] for item in items if item.get("skuCode")}),
                "warehouseCount": len({item["effectiveWarehouseSku"] for item in items if item.get("effectiveWarehouseSku")}),
                "skcCount": len({item["skcName"] for item in items if item.get("skcName")}),
                "spuCount": len({item["spuName"] for item in items if item.get("spuName")}),
                "mappedCount": sum(1 for item in items if item.get("mappingStatus") == "mapped"),
                "defaultCount": sum(1 for item in items if item.get("mappingStatus") == "default"),
                "unmappedCount": sum(1 for item in items if item.get("mappingStatus") == "unmapped"),
                "missingProductCount": sum(1 for item in items if item.get("mappingStatus") == "missing_product"),
                "missingSkuMappingCount": sum(1 for item in items if item.get("mappingStatus") in {"default", "unmapped"}),
                "inventoryTotal": sum(int(item.get("inventoryTotal") or 0) for item in items),
                "activeSkuCount": sum(1 for item in items if item.get("skcShelfStatus") == 1),
                "imageUrl": next((item.get("imageUrl") for item in items if item.get("imageUrl")), ""),
                "title": next((item.get("title") for item in items if item.get("title")), ""),
                "sampleSkus": [item["skuCode"] for item in items[:5] if item.get("skuCode")],
            }
        )
    return sorted(out, key=lambda item: (-item["skuCount"], item["key"]))


def warehouse_relation_rows_for_account(rows: list[dict[str, Any]], account: Account) -> list[dict[str, Any]]:
    if account.can(VIEW_WAREHOUSE_COST):
        return rows
    return [
        {key: value for key, value in row.items() if key not in {"costText", "costList"}}
        for row in rows
    ]

@app.get("/api/warehouse-relations")
def api_warehouse_relations(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> JSONResponse:
    account = require_auth(token, shein_pnl_token, permission=VIEW_WAREHOUSE_RELATIONS)
    url, shop_key = ensure_warehouse_relation_store()
    with psycopg.connect(url, row_factory=dict_row) as conn:
        default_row = conn.execute(
            "SELECT default_warehouse_sku FROM shein_sku_mapping_settings WHERE shop_key = %s",
            (shop_key,),
        ).fetchone()
        default_warehouse_sku = clean_text(default_row.get("default_warehouse_sku")) if default_row else ""
        mapping_rows = conn.execute(
            """
            SELECT id, warehouse_sku, shein_sku, sku_group, warehouse_qty, note, enabled, updated_at
            FROM shein_sku_mappings
            WHERE shop_key = %s AND enabled = true
            ORDER BY warehouse_sku, shein_sku, id
            """,
            (shop_key,),
        ).fetchall()
        warehouse_rows = conn.execute(
            """
            SELECT warehouse_sku, enabled, updated_at
            FROM shein_warehouse_skus
            WHERE shop_key = %s
            ORDER BY enabled DESC, warehouse_sku
            """,
            (shop_key,),
        ).fetchall()
        catalog = build_product_sku_catalog(conn, shop_key)

    rows: list[dict[str, Any]] = []
    mapped_skus: set[str] = set()
    for mapping in mapping_rows:
        sku_code = clean_text(mapping.get("shein_sku"))
        if not sku_code:
            continue
        mapped_skus.add(sku_code)
        rows.append(
            build_relation_row(
                shop_key=shop_key,
                sku_code=sku_code,
                product=catalog.get(sku_code),
                mapping=mapping,
                default_warehouse_sku=default_warehouse_sku,
            )
        )
    for sku_code, product in catalog.items():
        if sku_code in mapped_skus:
            continue
        rows.append(
            build_relation_row(
                shop_key=shop_key,
                sku_code=sku_code,
                product=product,
                mapping=None,
                default_warehouse_sku=default_warehouse_sku,
            )
        )

    product_sku_count = len(catalog)
    rows = warehouse_relation_rows_for_account(rows, account)

    summary = {
        "warehouseSkus": len({clean_text(row.get("warehouse_sku")) for row in warehouse_rows if clean_text(row.get("warehouse_sku"))}),
        "enabledWarehouseSkus": sum(1 for row in warehouse_rows if row.get("enabled")),
        "mappedRows": sum(1 for row in rows if row["mappingStatus"] == "mapped"),
        "productSkus": product_sku_count,
        "mappedProductSkus": len(mapped_skus.intersection(catalog.keys())),
        "unmappedProductSkus": len([row for row in rows if row["mappingStatus"] in {"default", "unmapped"}]),
        "missingSkuMappings": len([row for row in rows if row["mappingStatus"] in {"default", "unmapped"}]),
        "missingProductMappings": sum(1 for row in rows if row["mappingStatus"] == "missing_product"),
        "skcCount": len({row["skcName"] for row in rows if row.get("skcName")}),
        "spuCount": len({row["spuName"] for row in rows if row.get("spuName")}),
        "activeSkus": sum(1 for row in rows if row.get("skcShelfStatus") == 1),
        "offlineSkus": sum(1 for row in rows if row.get("skcShelfStatus") == 0),
        "inventoryTotal": sum(int(row.get("inventoryTotal") or 0) for row in rows),
        "defaultWarehouseSku": default_warehouse_sku,
        "coverageRate": round(len(mapped_skus.intersection(catalog.keys())) / product_sku_count, 4) if product_sku_count else None,
    }
    return JSONResponse(
        {
            "shop": shop_key,
            "summary": summary,
            "warehouseSkus": [
                {
                    "warehouseSku": clean_text(row.get("warehouse_sku")),
                    "enabled": bool(row.get("enabled")),
                    "updatedAt": serialize_relation_datetime(row.get("updated_at")),
                }
                for row in warehouse_rows
            ],
            "rows": rows,
            "groups": {
                "warehouses": relation_group_summary(rows, "effectiveWarehouseSku"),
                "spus": relation_group_summary(rows, "spuName"),
                "skcs": relation_group_summary(rows, "skcName"),
            },
        }
    )


@app.get("/api/sku-mappings")
def api_sku_mappings(q: str | None = None, token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_SKU_MAPPINGS)
    from .db import get_sku_mapping_settings, list_sku_mappings, list_warehouse_skus

    url, shop_key = ensure_sku_mapping_store()
    rows = list_sku_mappings(url, shop_key=shop_key, search=q)
    all_rows = list_sku_mappings(url, shop_key=shop_key)
    warehouse_skus = list_warehouse_skus(url, shop_key=shop_key)
    settings = get_sku_mapping_settings(url, shop_key=shop_key)
    default_warehouse_sku = clean_text(settings.get("default_warehouse_sku"))
    shein_skus = shein_sku_options()
    sku_meta = {item["sku"]: item for item in shein_skus}
    unmapped = unmapped_shein_sku_rows(shein_skus, all_rows, default_warehouse_sku)
    if q:
        q_text = q.strip().lower()
        unmapped = [row for row in unmapped if q_text in f"{row['warehouseSku']} {row['sheinSku']} {row['sheinSkuLabel']}".lower()]
    summary = summarize_sku_mappings(all_rows)
    summary["unmappedSheinSkus"] = len(unmapped_shein_sku_rows(shein_skus, all_rows, default_warehouse_sku))
    summary["defaultWarehouseSku"] = default_warehouse_sku
    summary["warehouseSkuCount"] = len(warehouse_skus)
    return JSONResponse({
        "shop": shop_key,
        "summary": summary,
        "settings": default_warehouse_sku_response(settings),
        "warehouseSkus": [serialize_warehouse_sku(row) for row in warehouse_skus],
        "sheinSkus": shein_skus,
        "rows": [serialize_sku_mapping(row, sku_meta=sku_meta) for row in rows],
        "unmappedRows": unmapped,
    })


@app.post("/api/warehouse-skus")
def api_save_warehouse_sku(
    payload: dict[str, Any] = Body(...),
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_SKU_MAPPINGS)
    from .db import upsert_warehouse_sku

    values = normalize_warehouse_sku_payload(payload)
    url, shop_key = ensure_sku_mapping_store()
    try:
        row, inserted = upsert_warehouse_sku(
            url,
            shop_key=shop_key,
            preserve_existing_values=not warehouse_detail_payload_present(payload),
            **values,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return JSONResponse({"row": serialize_warehouse_sku(row), "inserted": inserted})


@app.post("/api/sku-mappings/settings")
def api_save_sku_mapping_settings(
    payload: dict[str, Any] = Body(...),
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_SKU_MAPPINGS)
    from .db import save_sku_mapping_settings

    default_warehouse_sku = clean_text(payload_value(payload, "default_warehouse_sku", "defaultWarehouseSku"))
    if len(default_warehouse_sku) > 240:
        raise HTTPException(status_code=400, detail="default warehouse sku is too long")
    url, shop_key = ensure_sku_mapping_store()
    settings = save_sku_mapping_settings(url, shop_key=shop_key, default_warehouse_sku=default_warehouse_sku)
    return JSONResponse({"settings": default_warehouse_sku_response(settings)})


@app.post("/api/sku-mappings")
def api_save_sku_mapping(
    payload: dict[str, Any] = Body(...),
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_SKU_MAPPINGS)
    from .db import upsert_sku_mapping, upsert_warehouse_sku

    url, shop_key = ensure_sku_mapping_store()
    mapping_id_value = payload_value(payload, "id", "mappingId")
    mapping_id = int(mapping_id_value) if mapping_id_value not in (None, "") else None
    warehouse_values = normalize_warehouse_sku_payload(payload)
    values = normalize_sku_mapping_payload(payload)
    try:
        upsert_warehouse_sku(
            url,
            shop_key=shop_key,
            preserve_existing_values=not warehouse_detail_payload_present(payload),
            **warehouse_values,
        )
        row, inserted = upsert_sku_mapping(url, shop_key=shop_key, mapping_id=mapping_id, **values)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except psycopg.errors.UniqueViolation:
        raise HTTPException(status_code=409, detail="SKU mapping already exists")
    shein_skus = shein_sku_options()
    sku_meta = {item["sku"]: item for item in shein_skus}
    return JSONResponse({"row": serialize_sku_mapping(row, sku_meta=sku_meta), "inserted": inserted})


@app.delete("/api/sku-mappings/{mapping_id}")
def api_delete_sku_mapping(
    mapping_id: int,
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_SKU_MAPPINGS)
    from .db import delete_sku_mapping

    url, shop_key = ensure_sku_mapping_store()
    deleted = delete_sku_mapping(url, shop_key=shop_key, mapping_id=mapping_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="SKU mapping not found")
    return JSONResponse({"deleted": True})


@app.post("/api/sku-mappings/import")
async def api_import_sku_mappings(
    file: UploadFile = File(...),
    token: str | None = None,
    shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME),
) -> JSONResponse:
    require_auth(token, shein_pnl_token, permission=ACCESS_SKU_MAPPINGS)
    from .db import list_sku_mappings, upsert_sku_mapping, upsert_warehouse_sku

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Import file is too large")
    filename = file.filename or ""
    content_type = file.content_type or ""
    if filename.lower().endswith(".xls"):
        raise HTTPException(status_code=400, detail="Please save old .xls files as .xlsx before importing")
    records = parse_sku_mapping_excel(content) if is_excel_upload(filename, content_type) else parse_sku_mapping_csv(decode_csv_bytes(content))
    url, shop_key = ensure_sku_mapping_store()
    inserted = 0
    updated = 0
    skipped = 0
    errors: list[dict[str, Any]] = []
    for row_number, raw in records:
        try:
            warehouse_values = normalize_warehouse_sku_payload(raw)
            _, warehouse_inserted = upsert_warehouse_sku(
                url,
                shop_key=shop_key,
                preserve_existing_values=not warehouse_detail_payload_present(raw),
                **warehouse_values,
            )
            shein_sku = clean_text(raw.get("shein_sku"))
            if shein_sku:
                values = normalize_sku_mapping_payload(raw)
                _, mapping_inserted = upsert_sku_mapping(url, shop_key=shop_key, **values)
                is_inserted = warehouse_inserted or mapping_inserted
            else:
                is_inserted = warehouse_inserted
            inserted += 1 if is_inserted else 0
            updated += 0 if is_inserted else 1
        except HTTPException as exc:
            skipped += 1
            if len(errors) < 20:
                errors.append({"row": row_number, "error": exc.detail})
        except Exception as exc:
            skipped += 1
            if len(errors) < 20:
                errors.append({"row": row_number, "error": str(exc)})
    rows = list_sku_mappings(url, shop_key=shop_key)
    shein_skus = shein_sku_options()
    sku_meta = {item["sku"]: item for item in shein_skus}
    return JSONResponse({
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "summary": summarize_sku_mappings(rows),
        "rows": [serialize_sku_mapping(row, sku_meta=sku_meta) for row in rows],
    })


@app.get("/api/sku-mappings/export")
def api_export_sku_mappings(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> Response:
    require_auth(token, shein_pnl_token, permission=ACCESS_SKU_MAPPINGS)
    from .db import list_sku_mappings, list_warehouse_skus

    url, shop_key = ensure_sku_mapping_store()
    warehouse_rows = list_warehouse_skus(url, shop_key=shop_key, include_disabled=True)
    warehouse_by_sku = {clean_text(row.get("warehouse_sku")): row for row in warehouse_rows}
    mappings_by_warehouse: dict[str, list[dict[str, Any]]] = {}
    for row in list_sku_mappings(url, shop_key=shop_key):
        mappings_by_warehouse.setdefault(clean_text(row.get("warehouse_sku")), []).append(row)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(SKU_MAPPING_IMPORT_HEADERS)
    for warehouse_sku in sorted(set(warehouse_by_sku) | set(mappings_by_warehouse)):
        if not warehouse_sku:
            continue
        warehouse = warehouse_by_sku.get(warehouse_sku, {})
        mapping_rows = mappings_by_warehouse.get(warehouse_sku) or [{"shein_sku": ""}]
        for mapping in mapping_rows:
            writer.writerow([
                warehouse_sku,
                mapping.get("shein_sku") or "",
                warehouse.get("length_cm") if warehouse.get("length_cm") is not None else "",
                warehouse.get("width_cm") if warehouse.get("width_cm") is not None else "",
                warehouse.get("height_cm") if warehouse.get("height_cm") is not None else "",
                warehouse.get("weight_kg") if warehouse.get("weight_kg") is not None else "",
                warehouse.get("purchase_price") if warehouse.get("purchase_price") is not None else "",
                warehouse.get("ocean_freight_price") if warehouse.get("ocean_freight_price") is not None else "",
                warehouse.get("operation_fee_price") if warehouse.get("operation_fee_price") is not None else 0,
            ])
    return Response(
        "\ufeff" + output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="shein_sku_mappings.csv"'},
    )


@app.get("/api/sku-mappings/template")
def api_sku_mapping_template(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> Response:
    require_auth(token, shein_pnl_token, permission=ACCESS_SKU_MAPPINGS)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(SKU_MAPPING_IMPORT_HEADERS)
    writer.writerow(["WH-SKU-001", "SHEIN-SKU-001", "30", "20", "10", "0.5", "4.20", "0.80", "0"])
    writer.writerow(["WH-SKU-001", "SHEIN-SKU-002", "", "", "", "", "", "", ""])
    return Response(
        "\ufeff" + output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="shein_sku_mapping_template.csv"'},
    )


@app.get("/api/sku-mappings/template.xlsx")
def api_sku_mapping_excel_template(token: str | None = None, shein_pnl_token: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> Response:
    require_auth(token, shein_pnl_token, permission=ACCESS_SKU_MAPPINGS)
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError:
        raise HTTPException(status_code=500, detail="Excel template requires openpyxl. Run: pip install -r requirements.txt")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "sku_mappings"
    worksheet.append(SKU_MAPPING_IMPORT_HEADERS)
    worksheet.append(["WH-SKU-001", "SHEIN-SKU-001", "30", "20", "10", "0.5", "4.20", "0.80", "0"])
    worksheet.append(["WH-SKU-001", "SHEIN-SKU-002", "", "", "", "", "", "", ""])
    output = io.BytesIO()
    workbook.save(output)
    return Response(
        output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="shein_sku_mapping_template.xlsx"'},
    )


def serialize_row(row: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, value in row.items():
        if isinstance(value, pd.Timestamp):
            out[key] = value.strftime("%Y-%m-%d %H:%M")
        elif isinstance(value, (list, dict)):
            out[key] = value
        elif isinstance(value, float):
            out[key] = None if pd.isna(value) else round(float(value), 4)
        elif pd.isna(value):
            out[key] = None
        else:
            out[key] = value.item() if hasattr(value, "item") else value
    return out
